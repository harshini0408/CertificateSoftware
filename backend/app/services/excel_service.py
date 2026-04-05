"""Excel parsing utilities for participant import.

Expected Excel columns (case-insensitive, spaces/underscores normalised):
    - Email          (required — rows without email are skipped)
  - Role / Certificate Type / Type / cert_type
                   (optional — defaults to "participant" if missing)
  - Any other columns are preserved verbatim in the participant's ``fields`` dict.

The Role column value is lowercased and spaces replaced with underscores so
it matches the backend cert_type enum values:
  "Volunteer"  → "volunteer"
  "Winner 1st" → "winner_1st"
  "Coordinator"→ "coordinator"
  etc.
"""

from io import BytesIO
from typing import Any, Dict, List, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

# ── Column aliases for cert_type detection ────────────────────────────────────
_CERT_TYPE_ALIASES = {"role", "certificate type", "cert type", "type", "cert_type"}


def _normalise_key(k: str) -> str:
    """Lower-case + strip for alias matching."""
    return k.lower().strip()


def generate_excel_template(field_slots=None) -> BytesIO:
    """Create a downloadable .xlsx with standard certificate columns.

    Columns: Name, Email, Event Name, Event Date, Role
    - Role must be filled with the cert type (e.g. volunteer, winner, coordinator)
    - All other columns are the data that appears on the certificate
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Participants"

    headers = ["Name", "Registration Number", "Event Date", "Role", "Email"]
    sample   = ["John Doe", "21CS001", "2024-04-01", "participant", "john@example.com"]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)

    for col, value in enumerate(sample, 1):
        ws.cell(row=2, column=col, value=value)

    # Auto-width
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max(max_len + 4, 12)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def parse_participants_excel(
    file_bytes: bytes,
) -> Tuple[List[Dict[str, Any]], List[str]]:
    """Parse an uploaded Excel file and return (rows, errors).

    Expected columns: Name, Email, Event Name, Event Date, Role
    Each row dict contains all columns, plus ``_cert_type`` derived from Role.
    Rows missing Email are skipped with an error message.
    """
    wb = load_workbook(filename=BytesIO(file_bytes), read_only=True)
    ws = wb.active

    rows_iter = ws.iter_rows(values_only=True)
    headers_row = next(rows_iter, None)
    if not headers_row:
        return [], ["Excel file is empty"]

    headers = [str(h).strip() if h else "" for h in headers_row]
    # Map normalised key → original header index for cert_type detection
    cert_type_col_idx: int = -1
    for i, h in enumerate(headers):
        if _normalise_key(h) in _CERT_TYPE_ALIASES:
            cert_type_col_idx = i
            break

    parsed: List[Dict[str, Any]] = []
    errors: List[str] = []

    for row_idx, row in enumerate(rows_iter, start=2):
        if all(cell is None for cell in row):
            continue

        record: Dict[str, Any] = {}
        for col_idx, cell_value in enumerate(row):
            if col_idx < len(headers) and headers[col_idx]:
                record[headers[col_idx]] = str(cell_value).strip() if cell_value is not None else ""

        # Robustly find Email and Reg No regardless of exact column casing
        email = ""
        reg_no = ""
        for k, v in record.items():
            kl = k.lower().replace(" ", "").replace("_", "")
            if kl in ("email", "emailid", "mail"):
                email = v
            elif kl in ("registrationnumber", "regno", "rollno", "rollnumber"):
                reg_no = v

        if not email:
            errors.append(f"Row {row_idx}: missing Email — skipped")
            continue
            
        record["Email"] = email  # Enforce exact casing for the router


        # Derive cert_type from Role column (default to "participant")
        if cert_type_col_idx >= 0 and cert_type_col_idx < len(row):
            raw_role = str(row[cert_type_col_idx]).strip() if row[cert_type_col_idx] else ""
        else:
            raw_role = ""

        if raw_role:
            cert_type = raw_role.lower().replace(" ", "_").replace("-", "_")
        else:
            cert_type = "participant"

        # Store derived cert_type so the upload endpoint can use it
        record["_cert_type"] = cert_type

        parsed.append(record)

    wb.close()
    return parsed, errors
