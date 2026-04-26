import secrets
import string
import io
import re
from datetime import datetime
from typing import List, Optional

from beanie import PydanticObjectId
from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi import File, UploadFile
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field

from ...core.dependencies import require_role
from ...core.security import hash_password, verify_password
from ...models.user import User, UserRole
from ...models.club import Club
from ...models.department import Department
from ...models.event import Event
from ...models.certificate import Certificate, CertStatus
from ...models.dept_certificate import DeptCertificate
from ...models.dept_event import DeptEvent
from ...models.guest_session import GuestSession
from ...models.email_log import EmailLog, EmailStatus
from ...models.scan_log import ScanLog
from ...models.credit_rule import CreditRule
from ...models.student_credit import StudentCredit
from ...models.manual_credit_submission import ManualCreditSubmission, ManualSubmissionStatus
from ...schemas.club import ClubCreate, ClubUpdate, ClubResponse
from ...schemas.department import DepartmentCreate, DepartmentUpdate, DepartmentResponse
from ...schemas.credit import CreditRuleSchema, CreditRulesUpdateRequest, CreditRuleResponse
from ...schemas.user import UserCreate, UserUpdate, UserResponse

router = APIRouter(prefix="/admin", tags=["Admin"])

_admin = Depends(require_role(UserRole.SUPER_ADMIN))


class TutorStudentItem(BaseModel):
    name: str
    email: str
    registration_number: str


class TutorStudentAssignRequest(BaseModel):
    students: List[TutorStudentItem]


class TutorReassignRequest(BaseModel):
    new_tutor_id: PydanticObjectId


class CreditResetRequest(BaseModel):
    semester: str = Field(min_length=1, max_length=50)
    admin_password: str = Field(min_length=1, max_length=256)


async def _assign_student_to_tutor(
    *,
    tutor: User,
    name: str,
    email: str,
    registration_number: str,
) -> None:
    email_norm = email.strip().lower()
    reg_norm = registration_number.strip()
    if not email_norm or not reg_norm:
        raise ValueError("email and registration_number are required")

    existing = await StudentCredit.find_one(
        StudentCredit.student_email == email_norm,
    )

    if existing:
        updates = {
            "tutor_email": tutor.email,
            "last_updated": datetime.utcnow(),
        }
        if not existing.student_name and name.strip():
            updates["student_name"] = name.strip()
        if not existing.registration_number:
            updates["registration_number"] = reg_norm
        if not existing.department and tutor.department:
            updates["department"] = tutor.department
        if not existing.batch and tutor.batch:
            updates["batch"] = tutor.batch
        if not existing.section and tutor.section:
            updates["section"] = tutor.section
        await existing.set(updates)
        return

    await StudentCredit(
        student_email=email_norm,
        tutor_email=tutor.email,
        registration_number=reg_norm,
        student_name=name.strip(),
        department=tutor.department,
        batch=tutor.batch,
        section=tutor.section,
        total_credits=0,
        credit_history=[],
        last_updated=datetime.utcnow(),
    ).insert()


async def _find_tutor_by_email(email: str | None) -> User | None:
    email_norm = (email or "").strip().lower()
    if not email_norm:
        return None
    tutor = await User.find_one(User.email == email_norm)
    if not tutor or tutor.role != UserRole.TUTOR:
        return None
    return tutor


# ── Helper: Build UserResponse from User document ────────────────────────────

def _user_response(u: User) -> UserResponse:
    return UserResponse(
        id=str(u.id),
        username=u.username,
        name=u.name,
        email=u.email,
        role=u.role.value,
        is_active=u.is_active,
        created_at=u.created_at,
        club_id=str(u.club_id) if u.club_id else None,
        event_id=str(u.event_id) if u.event_id else None,
        department=u.department,
        registration_number=u.registration_number,
        batch=u.batch,
        section=u.section,
    )


def _club_response(c: Club) -> ClubResponse:
    return ClubResponse(
        id=str(c.id),
        name=c.name,
        slug=c.slug,
        contact_email=c.contact_email or "",
        is_active=c.is_active,
        created_at=c.created_at,
    )


def _department_response(d: Department) -> DepartmentResponse:
    slug_value = getattr(d, "slug", None) or _normalize_department_slug(getattr(d, "name", "")) or "DEPT"
    return DepartmentResponse(
        id=str(d.id),
        name=d.name,
        slug=slug_value,
        is_active=d.is_active,
        created_at=d.created_at,
    )


def _normalize_department_name(value: str | None) -> str:
    text = re.sub(r"\s+", " ", (value or "").strip())
    return text


def _normalize_department_slug(value: str | None) -> str:
    text = re.sub(r"[^A-Za-z0-9]", "", (value or "").strip())
    return text.upper()


async def _resolve_department_name(raw_value: str | None) -> str | None:
    raw_text = _normalize_department_name(raw_value)
    if not raw_text:
        return None

    normalized_slug = _normalize_department_slug(raw_text)
    escaped_slug = re.escape(normalized_slug)
    escaped_name = re.escape(raw_text)

    existing = await Department.find_one(
        {
            "$and": [
                {"is_active": True},
                {
                    "$or": [
                        {"slug": {"$regex": f"^{escaped_slug}$", "$options": "i"}},
                        {"name": {"$regex": f"^{escaped_name}$", "$options": "i"}},
                    ]
                },
            ]
        }
    )
    if not existing:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Department not found or inactive")
    return existing.name


def _normalize_cert_type(value: str | None) -> str:
    return (value or "").strip().lower().replace("-", "_").replace(" ", "_")


async def _find_credit_rule(cert_type_raw: str) -> Optional[CreditRule]:
    normalized = _normalize_cert_type(cert_type_raw)
    spaced = normalized.replace("_", " ")
    display = spaced.title() if normalized else cert_type_raw

    candidates = [cert_type_raw, normalized, spaced, display]
    seen = set()
    for candidate in candidates:
        key = (candidate or "").strip().lower()
        if not key or key in seen:
            continue
        seen.add(key)
        existing = await CreditRule.find_one({
            "cert_type": {"$regex": f"^{re.escape(candidate)}$", "$options": "i"}
        })
        if existing:
            return existing
    return None


# ═══ CLUBS ═══════════════════════════════════════════════════════════════════


@router.post("/clubs", response_model=ClubResponse, status_code=201)
async def create_club(body: ClubCreate, _user: User = _admin):
    slug_upper = body.slug.upper()

    if await Club.find_one(Club.slug == slug_upper):
        raise HTTPException(status.HTTP_409_CONFLICT, "Slug already exists")

    club = Club(
        name=body.name,
        slug=slug_upper,
    )
    await club.insert()

    return _club_response(club)


@router.get("/clubs", response_model=List[ClubResponse])
async def list_clubs(
    is_active: Optional[bool] = None,
    search: Optional[str] = None,
    _user: User = _admin,
):
    query = {}
    if is_active is not None:
        query["is_active"] = is_active
    if search:
        import re
        pattern = re.compile(re.escape(search), re.IGNORECASE)
        query["$or"] = [{"name": pattern}, {"slug": pattern}]

    clubs = await Club.find(query).to_list()
    return [_club_response(c) for c in clubs]


@router.get("/clubs/{club_id}", response_model=ClubResponse)
async def get_club(club_id: PydanticObjectId, _user: User = _admin):
    club = await Club.get(club_id)
    if not club:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Club not found")
    return _club_response(club)


@router.patch("/clubs/{club_id}", response_model=ClubResponse)
async def update_club(
    club_id: PydanticObjectId,
    body: ClubUpdate,
    _user: User = _admin,
):
    club = await Club.get(club_id)
    if not club:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Club not found")

    update_data = body.model_dump(exclude_none=True)

    # If deactivating club, cascade to all users belonging to this club
    if update_data.get("is_active") is False and club.is_active is True:
        await User.find(
            User.club_id == club_id,
            User.is_active == True,
        ).update_many({"$set": {"is_active": False}})

    if update_data:
        await club.set(update_data)

    return _club_response(club)


@router.patch("/clubs/{club_id}/toggle-active", response_model=ClubResponse)
async def toggle_club_active(club_id: PydanticObjectId, _user: User = _admin):
    club = await Club.get(club_id)
    if not club:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Club not found")

    club.is_active = not club.is_active
    if club.is_active is False:
        await User.find(User.club_id == club_id).update_many({"$set": {"is_active": False}})

    await club.save()
    return _club_response(club)


@router.get("/clubs/{club_id}/users", response_model=List[UserResponse])
async def get_club_users(club_id: PydanticObjectId, _user: User = _admin):
    club = await Club.get(club_id)
    if not club:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Club not found")

    users = await User.find(User.club_id == club_id).to_list()
    return [_user_response(u) for u in users]


# ═══ DEPARTMENTS ═════════════════════════════════════════════════════════════


@router.post("/departments", response_model=DepartmentResponse, status_code=201)
async def create_department(body: DepartmentCreate, _user: User = _admin):
    normalized_name = _normalize_department_name(body.name)
    normalized_slug = _normalize_department_slug(body.slug)
    if not normalized_name:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Department name is required")
    if not normalized_slug:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Department slug is required")
    if not re.fullmatch(r"[A-Z0-9]{2,20}", normalized_slug):
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Department slug must use uppercase letters and digits only")

    escaped_name = re.escape(normalized_name)
    escaped_slug = re.escape(normalized_slug)
    existing = await Department.find_one({
        "$or": [
            {"name": {"$regex": f"^{escaped_name}$", "$options": "i"}},
            {"slug": {"$regex": f"^{escaped_slug}$", "$options": "i"}},
        ]
    })
    if existing and existing.name.lower() == normalized_name.lower():
        raise HTTPException(status.HTTP_409_CONFLICT, "Department already exists")
    if existing:
        raise HTTPException(status.HTTP_409_CONFLICT, "Department slug already exists")

    department = Department(name=normalized_name, slug=normalized_slug)
    await department.insert()
    return _department_response(department)


@router.get("/departments", response_model=List[DepartmentResponse])
async def list_departments(
    is_active: Optional[bool] = None,
    search: Optional[str] = None,
    _user: User = _admin,
):
    query = {}
    if is_active is not None:
        query["is_active"] = is_active
    if search:
        pattern = re.compile(re.escape(search), re.IGNORECASE)
        query["$or"] = [{"name": pattern}, {"slug": pattern}]

    departments = await Department.find(query).sort("name").to_list()

    # Backfill slug for legacy records that were created before slug support.
    for d in departments:
        if getattr(d, "slug", None):
            continue

        base_slug = _normalize_department_slug(getattr(d, "name", "")) or "DEPT"
        slug_candidate = base_slug[:20]
        suffix = 1
        while True:
            duplicate = await Department.find_one({"slug": slug_candidate})
            if not duplicate or duplicate.id == d.id:
                break
            suffix_str = str(suffix)
            keep = max(1, 20 - len(suffix_str))
            slug_candidate = f"{base_slug[:keep]}{suffix_str}"
            suffix += 1

        await d.set({"slug": slug_candidate})

    if departments:
        departments = await Department.find(query).sort("name").to_list()

    return [_department_response(d) for d in departments]


@router.patch("/departments/{department_id}", response_model=DepartmentResponse)
async def update_department(
    department_id: PydanticObjectId,
    body: DepartmentUpdate,
    _user: User = _admin,
):
    department = await Department.get(department_id)
    if not department:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Department not found")

    update_data = body.model_dump(exclude_none=True)

    if "name" in update_data:
        normalized_name = _normalize_department_name(update_data["name"])
        if not normalized_name:
            raise HTTPException(status.HTTP_400_BAD_REQUEST, "Department name is required")

        escaped = re.escape(normalized_name)
        duplicate = await Department.find_one({"name": {"$regex": f"^{escaped}$", "$options": "i"}})
        if duplicate and duplicate.id != department.id:
            raise HTTPException(status.HTTP_409_CONFLICT, "Department already exists")

        old_name = department.name
        if normalized_name != old_name:
            update_data["name"] = normalized_name
            await User.find(User.department == old_name).update_many({"$set": {"department": normalized_name}})

    if "slug" in update_data:
        normalized_slug = _normalize_department_slug(update_data["slug"])
        if not normalized_slug:
            raise HTTPException(status.HTTP_400_BAD_REQUEST, "Department slug is required")
        if not re.fullmatch(r"[A-Z0-9]{2,20}", normalized_slug):
            raise HTTPException(status.HTTP_400_BAD_REQUEST, "Department slug must use uppercase letters and digits only")

        escaped_slug = re.escape(normalized_slug)
        duplicate_slug = await Department.find_one({"slug": {"$regex": f"^{escaped_slug}$", "$options": "i"}})
        if duplicate_slug and duplicate_slug.id != department.id:
            raise HTTPException(status.HTTP_409_CONFLICT, "Department slug already exists")

        update_data["slug"] = normalized_slug

    if update_data:
        await department.set(update_data)

    updated = await Department.get(department_id)
    if not updated:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Department not found")
    return _department_response(updated)


@router.delete("/departments/{department_id}")
async def delete_department(department_id: PydanticObjectId, _user: User = _admin):
    department = await Department.get(department_id)
    if not department:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Department not found")

    if await User.find_one(User.department == department.name):
        raise HTTPException(
            status.HTTP_409_CONFLICT,
            "Department is assigned to existing users. Reassign them before deleting.",
        )

    await department.delete()
    return {"message": "Department deleted"}


# ═══ USERS ═══════════════════════════════════════════════════════════════════


@router.post("/users", response_model=UserResponse, status_code=201)
async def create_user(body: UserCreate, _user: User = _admin):
    username = body.username.strip()
    email = body.email.strip().lower()
    name = body.name.strip()
    password_input = body.password.strip()

    if len(password_input) < 8:
        raise HTTPException(status.HTTP_422_UNPROCESSABLE_ENTITY, "Password must be at least 8 characters")

    # Uniqueness checks
    if await User.find_one(User.username == username):
        raise HTTPException(status.HTTP_409_CONFLICT, "Username already taken")
    if await User.find_one(User.email == email):
        raise HTTPException(status.HTTP_409_CONFLICT, "Email already in use")

    # Club validation
    club_oid = None
    if body.club_id and body.role != "guest":
        club_oid = PydanticObjectId(body.club_id)
        club = await Club.get(club_oid)
        if not club or not club.is_active:
            raise HTTPException(
                status.HTTP_404_NOT_FOUND,
                "Club not found or inactive",
            )

    # Event validation
    event_oid = None
    if body.event_id and body.role != "guest":
        event_oid = PydanticObjectId(body.event_id)
        event = await Event.get(event_oid)
        if not event:
            raise HTTPException(status.HTTP_404_NOT_FOUND, "Event not found")
        if club_oid and event.club_id != club_oid:
            raise HTTPException(
                status.HTTP_404_NOT_FOUND,
                "Event does not belong to the specified club",
            )

    # Student registration_number uniqueness
    if body.role == "student" and body.registration_number:
        if await User.find_one(
            User.registration_number == body.registration_number
        ):
            raise HTTPException(
                status.HTTP_409_CONFLICT,
                "Registration number already registered",
            )

    department_name = None
    if body.role in ["dept_coordinator", "hod", "student", "tutor"]:
        department_name = await _resolve_department_name(body.department)
        if not department_name:
            raise HTTPException(status.HTTP_400_BAD_REQUEST, "Department is required for this role")

    # Enforce only one HOD account per department.
    if body.role == "hod" and department_name:
        existing_hod = await User.find_one(
            User.role == UserRole.HOD,
            User.department == department_name,
        )
        if existing_hod:
            raise HTTPException(
                status.HTTP_409_CONFLICT,
                f"HOD already exists for department '{department_name}'",
            )

    new_user = User(
        username=username,
        name=name,
        email=email,
        password_hash=hash_password(password_input),
        role=UserRole(body.role),
        first_login_completed=(body.role != "club_coordinator"),
        is_active=body.is_active,
        club_id=club_oid,
        event_id=event_oid,
        department=department_name,
        registration_number=body.registration_number.strip() if body.registration_number and body.role not in ["guest", "club_coordinator", "dept_coordinator", "tutor"] else None,
        batch=body.batch.strip() if body.batch and body.role not in ["guest", "club_coordinator", "dept_coordinator", "hod"] else None,
        section=body.section.strip() if body.section and body.role not in ["guest", "club_coordinator", "dept_coordinator", "hod"] else None,
    )
    await new_user.insert()

    # Auto-create student_credits doc for students
    if new_user.role == UserRole.STUDENT and body.registration_number:
        existing_credit = await StudentCredit.find_one(
            StudentCredit.student_email == email,
        )
        if not existing_credit:
            await StudentCredit(
                student_email=email,
                registration_number=body.registration_number.strip() if body.registration_number else body.registration_number,
                student_name=name,
                department=department_name,
                batch=body.batch.strip() if body.batch else body.batch,
                section=body.section.strip() if body.section else body.section,
                total_credits=0,
                credit_history=[],
                last_updated=datetime.utcnow(),
            ).insert()

    return _user_response(new_user)


@router.post("/users/bulk-import", status_code=200)
async def bulk_import_students(
    file: UploadFile = File(...),
    _user: User = _admin,
):
    """Import multiple students from an .xlsx file.

        Expected columns (case-insensitive, in any order):
            name, email, username, password, department, registration_number, batch, section
        Optional:
            tutor_email  (used to map imported students to tutor accounts)

    Returns: { created: int, skipped: int, errors: [ { row: int, reason: str } ] }
    """
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Only .xlsx files are accepted")

    import openpyxl
    contents = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
        ws = wb.active
    except Exception:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Could not parse the Excel file. Ensure it is a valid .xlsx.")

    # Read headers from row 1, normalize to lowercase
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Excel file is empty or has no header row")
    headers = [str(h).strip().lower() if h is not None else "" for h in header_row]

    REQUIRED = {"name", "email", "username", "password", "department", "registration_number", "batch", "section"}
    missing = REQUIRED - set(headers)
    if missing:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            f"Missing required columns: {', '.join(sorted(missing))}. Found: {', '.join(h for h in headers if h)}"
        )

    def col(row_vals, field):
        try:
            idx = headers.index(field)
            v = row_vals[idx]
            return str(v).strip() if v is not None else ""
        except (ValueError, IndexError):
            return ""

    created = 0
    skipped = 0
    errors = []

    for row_idx, row_vals in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # Skip fully empty rows
        if all(v is None or str(v).strip() == "" for v in row_vals):
            continue

        name      = col(row_vals, "name")
        email     = col(row_vals, "email").lower()
        username  = col(row_vals, "username")
        password  = col(row_vals, "password")
        dept      = col(row_vals, "department")
        reg_no    = col(row_vals, "registration_number")
        batch     = col(row_vals, "batch")
        section   = col(row_vals, "section")
        tutor_email = col(row_vals, "tutor_email").lower()

        try:
            # Validate required fields
            missing_fields = [f for f, v in [
                ("name", name), ("email", email), ("username", username),
                ("password", password), ("department", dept),
                ("registration_number", reg_no), ("batch", batch), ("section", section)
            ] if not v]
            if missing_fields:
                raise ValueError(f"Missing: {', '.join(missing_fields)}")

            if len(password) < 8:
                raise ValueError("Password must be at least 8 characters")

            resolved_dept = await _resolve_department_name(dept)
            if not resolved_dept:
                raise ValueError("Department is required")

            tutor = None
            if tutor_email:
                tutor = await _find_tutor_by_email(tutor_email)
                if not tutor:
                    raise ValueError(f"Tutor email '{tutor_email}' not found")

            # Uniqueness checks
            if await User.find_one(User.username == username):
                skipped += 1
                errors.append({"row": row_idx, "reason": f"Username '{username}' already exists — skipped"})
                continue
            if await User.find_one(User.email == email):
                skipped += 1
                errors.append({"row": row_idx, "reason": f"Email '{email}' already exists — skipped"})
                continue
            if await User.find_one(User.registration_number == reg_no):
                skipped += 1
                errors.append({"row": row_idx, "reason": f"Registration number '{reg_no}' already exists — skipped"})
                continue

            # Create User
            new_user = User(
                username=username,
                name=name,
                email=email,
                password_hash=hash_password(password),
                role=UserRole.STUDENT,
                is_active=True,
                department=resolved_dept,
                registration_number=reg_no,
                batch=batch,
                section=section,
            )
            await new_user.insert()

            # Create StudentCredit doc
            existing_credit = await StudentCredit.find_one(StudentCredit.student_email == email)
            if not existing_credit:
                await StudentCredit(
                    student_email=email,
                    tutor_email=(tutor.email if tutor else None),
                    registration_number=reg_no,
                    student_name=name,
                    department=resolved_dept,
                    batch=batch,
                    section=section,
                    total_credits=0,
                    credit_history=[],
                    last_updated=datetime.utcnow(),
                ).insert()
            elif tutor and existing_credit.tutor_email != tutor.email:
                await existing_credit.set({"tutor_email": tutor.email, "last_updated": datetime.utcnow()})

            created += 1

        except ValueError as ve:
            errors.append({"row": row_idx, "reason": str(ve)})
        except Exception as exc:
            errors.append({"row": row_idx, "reason": f"Unexpected error: {exc}"})

    return {"created": created, "skipped": skipped, "errors": errors}


@router.get("/users/bulk-import-tutors/sample")
async def download_tutor_import_sample(_user: User = _admin):
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tutors"

    headers = ["name", "username", "email", "password", "department", "batch", "section"]
    ws.append(headers)
    ws.append(["Jane Tutor", "jane_tutor", "jane.tutor@example.com", "Passw0rd!", "CSE", "2024-2028", "A"])

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)

    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=tutor_bulk_import_sample.xlsx"},
    )


@router.post("/users/bulk-import-tutors", status_code=200)
async def bulk_import_tutors(
    file: UploadFile = File(...),
    _user: User = _admin,
):
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Only .xlsx files are accepted")

    import openpyxl
    contents = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
        ws = wb.active
    except Exception:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Could not parse the Excel file. Ensure it is a valid .xlsx.")

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Excel file is empty or has no header row")

    headers = [str(h).strip().lower() if h is not None else "" for h in header_row]
    required = {"name", "username", "email", "password", "department", "batch", "section"}
    missing = required - set(headers)
    if missing:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            f"Missing required columns: {', '.join(sorted(missing))}. Found: {', '.join(h for h in headers if h)}",
        )

    def col(row_vals, field):
        try:
            idx = headers.index(field)
            v = row_vals[idx]
            return str(v).strip() if v is not None else ""
        except (ValueError, IndexError):
            return ""

    created = 0
    skipped = 0
    errors = []

    for row_idx, row_vals in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(v is None or str(v).strip() == "" for v in row_vals):
            continue

        name = col(row_vals, "name")
        username = col(row_vals, "username")
        email = col(row_vals, "email").lower()
        password = col(row_vals, "password")
        department = col(row_vals, "department")
        batch = col(row_vals, "batch")
        section = col(row_vals, "section")

        try:
            missing_fields = [
                f for f, v in [
                    ("name", name),
                    ("username", username),
                    ("email", email),
                    ("password", password),
                    ("department", department),
                    ("batch", batch),
                    ("section", section),
                ] if not v
            ]
            if missing_fields:
                raise ValueError(f"Missing: {', '.join(missing_fields)}")

            if len(password) < 8:
                raise ValueError("Password must be at least 8 characters")

            resolved_dept = await _resolve_department_name(department)
            if not resolved_dept:
                raise ValueError("Department is required")

            if await User.find_one(User.username == username):
                skipped += 1
                errors.append({"row": row_idx, "reason": f"Username '{username}' already exists — skipped"})
                continue

            if await User.find_one(User.email == email):
                skipped += 1
                errors.append({"row": row_idx, "reason": f"Email '{email}' already exists — skipped"})
                continue

            tutor = User(
                username=username,
                name=name,
                email=email,
                password_hash=hash_password(password),
                role=UserRole.TUTOR,
                first_login_completed=True,
                is_active=True,
                department=resolved_dept,
                batch=batch,
                section=section,
            )
            await tutor.insert()
            created += 1

        except ValueError as ve:
            errors.append({"row": row_idx, "reason": str(ve)})
        except Exception as exc:
            errors.append({"row": row_idx, "reason": f"Unexpected error: {exc}"})

    return {"created": created, "skipped": skipped, "errors": errors}


@router.post("/tutors/{tutor_id}/students")
async def assign_tutor_students(
    tutor_id: PydanticObjectId,
    body: TutorStudentAssignRequest,
    _user: User = _admin,
):
    tutor = await User.get(tutor_id)
    if not tutor or tutor.role != UserRole.TUTOR:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Tutor not found")

    created_or_updated = 0
    errors = []
    for idx, item in enumerate(body.students, start=1):
        try:
            await _assign_student_to_tutor(
                tutor=tutor,
                name=item.name,
                email=item.email,
                registration_number=item.registration_number,
            )
            created_or_updated += 1
        except Exception as exc:
            errors.append({"row": idx, "reason": str(exc)})

    return {
        "assigned": created_or_updated,
        "errors": errors,
    }


@router.post("/tutors/{tutor_id}/students/import")
async def bulk_import_tutor_students(
    tutor_id: PydanticObjectId,
    file: UploadFile = File(...),
    _user: User = _admin,
):
    tutor = await User.get(tutor_id)
    if not tutor or tutor.role != UserRole.TUTOR:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Tutor not found")

    if not file.filename.endswith(".xlsx"):
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Only .xlsx files are accepted")

    import openpyxl
    contents = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
        ws = wb.active
    except Exception:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Could not parse the Excel file. Ensure it is a valid .xlsx.")

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Excel file is empty or has no header row")

    raw_headers = [str(h).strip() if h is not None else "" for h in header_row]
    norm_headers = [re.sub(r"[^a-z0-9]", "", h.lower()) for h in raw_headers]

    def idx_for(*aliases: str) -> int:
        for alias in aliases:
            alias_norm = re.sub(r"[^a-z0-9]", "", alias.lower())
            if alias_norm in norm_headers:
                return norm_headers.index(alias_norm)
        return -1

    name_idx = idx_for("name", "student_name")
    email_idx = idx_for("email", "student_email")
    reg_idx = idx_for("registration number", "registration_number", "reg_no", "registrationno")

    if name_idx < 0 or email_idx < 0 or reg_idx < 0:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            "Missing required columns: name, email, registration number",
        )

    assigned = 0
    skipped = 0
    errors = []

    for row_idx, row_vals in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(v is None or str(v).strip() == "" for v in row_vals):
            continue

        name = str(row_vals[name_idx]).strip() if name_idx < len(row_vals) and row_vals[name_idx] is not None else ""
        email = str(row_vals[email_idx]).strip().lower() if email_idx < len(row_vals) and row_vals[email_idx] is not None else ""
        reg_no = str(row_vals[reg_idx]).strip() if reg_idx < len(row_vals) and row_vals[reg_idx] is not None else ""

        if not name or not email or not reg_no:
            skipped += 1
            errors.append({"row": row_idx, "reason": "Missing name/email/registration number"})
            continue

        try:
            await _assign_student_to_tutor(
                tutor=tutor,
                name=name,
                email=email,
                registration_number=reg_no,
            )
            assigned += 1
        except Exception as exc:
            skipped += 1
            errors.append({"row": row_idx, "reason": str(exc)})

    return {
        "assigned": assigned,
        "skipped": skipped,
        "errors": errors,
    }


@router.post("/tutors/{tutor_id}/reassign")
async def reassign_tutor_students(
    tutor_id: PydanticObjectId,
    body: TutorReassignRequest,
    _user: User = _admin,
):
    from_tutor = await User.get(tutor_id)
    if not from_tutor or from_tutor.role != UserRole.TUTOR:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Source tutor not found")

    to_tutor = await User.get(body.new_tutor_id)
    if not to_tutor or to_tutor.role != UserRole.TUTOR:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Target tutor not found")

    if from_tutor.id == to_tutor.id:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Source and target tutor cannot be the same")

    mapped = await StudentCredit.find(StudentCredit.tutor_email == from_tutor.email).to_list()
    moved = 0
    now = datetime.utcnow()
    for doc in mapped:
        await doc.set(
            {
                "tutor_email": to_tutor.email,
                "department": to_tutor.department,
                "batch": to_tutor.batch,
                "section": to_tutor.section,
                "last_updated": now,
            }
        )
        moved += 1

    return {
        "moved": moved,
        "from_tutor": from_tutor.email,
        "to_tutor": to_tutor.email,
    }


@router.get("/users", response_model=List[UserResponse])
async def list_users(
    role: Optional[str] = None,
    club_id: Optional[str] = None,
    department: Optional[str] = None,
    is_active: Optional[bool] = None,
    search: Optional[str] = None,
    _user: User = _admin,
):
    query = {}
    if role:
        query["role"] = role
    if club_id:
        query["club_id"] = PydanticObjectId(club_id)
    if department:
        query["department"] = department
    if is_active is not None:
        query["is_active"] = is_active
    if search:
        import re
        pattern = re.compile(re.escape(search), re.IGNORECASE)
        query["$or"] = [
            {"name": pattern},
            {"username": pattern},
            {"email": pattern},
        ]

    users = await User.find(query).to_list()
    return [_user_response(u) for u in users]


@router.get("/student-certificates/search")
async def search_student_certificates(
    q: str = Query(..., min_length=1),
    _user: User = _admin,
):
    query_text = q.strip()
    if not query_text:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Search query is required")

    pattern = re.compile(re.escape(query_text), re.IGNORECASE)

    students = await User.find({
        "role": UserRole.STUDENT,
        "$or": [
            {"name": pattern},
            {"email": pattern},
            {"registration_number": pattern},
        ],
    }).to_list()

    if not students:
        credits = await StudentCredit.find({
            "$or": [
                {"student_email": {"$regex": f"^{re.escape(query_text.lower())}$", "$options": "i"}},
                {"registration_number": query_text},
            ]
        }).to_list()
        emails = {c.student_email for c in credits if c.student_email}
        if emails:
            students = await User.find({"role": UserRole.STUDENT, "email": {"$in": list(emails)}}).to_list()

    results = []
    for student in students:
        email = (student.email or "").strip().lower()
        reg_no = (student.registration_number or "").strip()

        generated_certs = await Certificate.find({
            "snapshot.email": {"$regex": f"^{re.escape(email)}$", "$options": "i"},
        }).sort("-issued_at").to_list()

        manual_submissions = await ManualCreditSubmission.find({
            "student_email": {"$regex": f"^{re.escape(email)}$", "$options": "i"},
        }).sort("-submitted_at").to_list()

        credit_doc = await StudentCredit.find_one({
            "$or": [
                {"student_email": {"$regex": f"^{re.escape(email)}$", "$options": "i"}},
                {"registration_number": reg_no} if reg_no else {"_id": None},
            ]
        })

        credit_points_by_cert: dict[str, int] = {}
        if credit_doc:
            for entry in credit_doc.credit_history or []:
                if entry.cert_number:
                    credit_points_by_cert[entry.cert_number] = int(entry.points_awarded or 0)

        certificates = []
        for cert in generated_certs:
            certificates.append({
                "source_type": "generated",
                "cert_number": cert.cert_number,
                "event_name": cert.snapshot.event_name if cert.snapshot else "",
                "club_name": cert.snapshot.club_name if cert.snapshot else "",
                "status": cert.status.value,
                "issued_at": cert.issued_at,
                "credit_points": int(credit_points_by_cert.get(cert.cert_number, 0)),
            })

        for submission in manual_submissions:
            cert_number = f"STU-MANUAL-{str(submission.id)[-8:].upper()}"
            certificates.append({
                "source_type": "manual_upload",
                "cert_number": cert_number,
                "event_name": "Student Manual Submission",
                "club_name": "Student Upload",
                "status": submission.status.value,
                "issued_at": submission.submitted_at,
                "credit_points": int(submission.points_awarded or 0),
            })

        certificates.sort(key=lambda item: item.get("issued_at") or datetime.min, reverse=True)

        results.append({
            "student": _user_response(student),
            "total_certificates": len(certificates),
            "generated_count": len(generated_certs),
            "manual_upload_count": len(manual_submissions),
            "verified_manual_count": sum(1 for s in manual_submissions if s.status == ManualSubmissionStatus.VERIFIED),
            "certificates": certificates,
        })

    return {
        "query": query_text,
        "count": len(results),
        "results": results,
    }


@router.get("/tutors/mapping-summary")
async def tutor_mapping_summary(
    _user: User = _admin,
):
    tutors = await User.find(User.role == UserRole.TUTOR).to_list()
    credits = await StudentCredit.find_all().to_list()

    mapped_by_tutor = {}
    for credit in credits:
        email = (credit.tutor_email or "").strip().lower()
        if not email:
            continue
        mapped_by_tutor[email] = mapped_by_tutor.get(email, 0) + 1

    items = []
    total_mapped = 0
    for tutor in tutors:
        tutor_email = (tutor.email or "").strip().lower()
        mapped_count = int(mapped_by_tutor.get(tutor_email, 0))
        total_mapped += mapped_count
        items.append(
            {
                "id": str(tutor.id),
                "name": tutor.name,
                "email": tutor.email,
                "department": tutor.department,
                "batch": tutor.batch,
                "section": tutor.section,
                "mapped_students": mapped_count,
            }
        )

    items.sort(key=lambda x: x["mapped_students"], reverse=True)

    return {
        "total_tutors": len(tutors),
        "total_mapped_students": total_mapped,
        "items": items,
    }


@router.get("/users/{user_id}", response_model=UserResponse)
async def get_user(user_id: PydanticObjectId, _user: User = _admin):
    target = await User.get(user_id)
    if not target:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "User not found")
    return _user_response(target)


@router.patch("/users/{user_id}", response_model=UserResponse)
async def update_user(
    user_id: PydanticObjectId,
    body: UserUpdate,
    _user: User = _admin,
):
    target = await User.get(user_id)
    if not target:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "User not found")

    updates = body.model_dump(exclude_none=True)

    # Username uniqueness check
    if "username" in updates and updates["username"] != target.username:
        if await User.find_one(User.username == updates["username"]):
            raise HTTPException(status.HTTP_409_CONFLICT, "Username already taken")

    # Email uniqueness check
    if "email" in updates and updates["email"] != target.email:
        if await User.find_one(User.email == updates["email"]):
            raise HTTPException(status.HTTP_409_CONFLICT, "Email already in use")

    if updates:
        await target.set(updates)

    return _user_response(target)


@router.patch("/users/{user_id}/toggle-active", response_model=UserResponse)
async def toggle_user_active(user_id: PydanticObjectId, _user: User = _admin):
    target = await User.get(user_id)
    if not target:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "User not found")
    if target.role == UserRole.SUPER_ADMIN:
        raise HTTPException(status.HTTP_403_FORBIDDEN, "Cannot toggle the super admin account")

    target.is_active = not target.is_active
    await target.save()
    return _user_response(target)


@router.post("/users/{user_id}/reset-password")
async def reset_user_password(user_id: PydanticObjectId, _user: User = _admin):
    target = await User.get(user_id)
    if not target:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "User not found")

    tmp_pw = "".join(secrets.choice(string.ascii_letters + string.digits) for _ in range(12))
    target.password_hash = hash_password(tmp_pw)
    await target.save()

    return {"message": "Password reset successfully", "temp_password": tmp_pw}


@router.delete("/users/{user_id}")
async def deactivate_user(user_id: PydanticObjectId, _user: User = _admin):
    target = await User.get(user_id)
    if not target:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "User not found")
    if target.role == UserRole.SUPER_ADMIN:
        raise HTTPException(
            status.HTTP_403_FORBIDDEN,
            "Cannot deactivate the super admin account",
        )

    await target.set({"is_active": False})
    return {"message": "User deactivated successfully"}


# ═══ CERTIFICATES ════════════════════════════════════════════════════════════


@router.get("/certificates")
async def list_certificates(
    club_id: Optional[str] = None,
    event_id: Optional[str] = None,
    cert_status: Optional[CertStatus] = Query(None, alias="status"),
    date_from: Optional[datetime] = None,
    date_to: Optional[datetime] = None,
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
    _user: User = _admin,
):
    query = {}
    if club_id:
        query["club_id"] = PydanticObjectId(club_id)
    if event_id:
        query["event_id"] = PydanticObjectId(event_id)
    if cert_status:
        query["status"] = cert_status.value
    if date_from or date_to:
        date_q = {}
        if date_from:
            date_q["$gte"] = date_from
        if date_to:
            date_q["$lte"] = date_to
        query["issued_at"] = date_q

    skip = (page - 1) * page_size
    certs = await Certificate.find(query).skip(skip).limit(page_size).to_list()
    total = await Certificate.find(query).count()
    return {"total": total, "page": page, "page_size": page_size, "items": [
        {"id": str(c.id), "cert_number": c.cert_number, "status": c.status.value,
         "snapshot": c.snapshot.model_dump(), "issued_at": c.issued_at}
        for c in certs
    ]}


@router.patch("/certificates/{cert_number}/revoke")
async def revoke_certificate(cert_number: str, admin: User = _admin):
    cert = await Certificate.find_one(Certificate.cert_number == cert_number)
    if not cert:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Certificate not found")

    if cert.status == CertStatus.EMAILED:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            "Emailed certificates cannot be revoked",
        )

    if cert.status == CertStatus.REVOKED:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Certificate is already revoked")

    await cert.set({
        "status": CertStatus.REVOKED,
        "revoked_at": datetime.utcnow(),
        "revoked_by": admin.id,
    })
    return {"message": f"Certificate {cert_number} revoked"}


# ═══ SCAN LOGS ═══════════════════════════════════════════════════════════════


@router.get("/scan-logs")
async def list_scan_logs(
    cert_number: Optional[str] = None,
    date_from: Optional[datetime] = None,
    date_to: Optional[datetime] = None,
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
    _user: User = _admin,
):
    query = {}
    if cert_number:
        query["cert_number"] = cert_number
    if date_from or date_to:
        date_q = {}
        if date_from:
            date_q["$gte"] = date_from
        if date_to:
            date_q["$lte"] = date_to
        query["scanned_at"] = date_q

    skip = (page - 1) * page_size
    logs = await ScanLog.find(query).skip(skip).limit(page_size).to_list()
    total = await ScanLog.find(query).count()
    return {"total": total, "page": page, "items": [
        {"id": str(l.id), "cert_number": l.cert_number, "ip_address": l.ip_address,
         "user_agent": l.user_agent, "scanned_at": l.scanned_at}
        for l in logs
    ]}


@router.get("/activity")
async def recent_activity(_user: User = _admin):
    certs = await Certificate.find_all().sort("-issued_at").limit(20).to_list()

    return [
        {
            "action": "Certificate Issued",
            "actor": cert.snapshot.name if cert.snapshot else "Unknown",
            "target": cert.snapshot.event_name if cert.snapshot else "",
            "cert_number": cert.cert_number,
            "timestamp": cert.issued_at,
            "club_name": cert.snapshot.club_name if cert.snapshot else "",
        }
        for cert in certs
    ]


# ═══ CREDIT RULES ════════════════════════════════════════════════════════════


@router.get("/credit-rules", response_model=List[CreditRuleResponse])
async def get_credit_rules(_user: User = _admin):
    rules = await CreditRule.find_all().to_list()
    return [CreditRuleResponse(id=str(r.id), cert_type=r.cert_type, points=r.points,
                               updated_by=str(r.updated_by) if r.updated_by else None,
                               updated_at=r.updated_at) for r in rules]


@router.put("/credit-rules")
async def upsert_credit_rules(body: CreditRulesUpdateRequest, admin: User = _admin):
    for rule in body.rules:
        existing = await _find_credit_rule(rule.cert_type)
        if existing:
            await existing.set({"cert_type": rule.cert_type, "points": rule.points, "updated_by": admin.id,
                                "updated_at": datetime.utcnow()})
        else:
            await CreditRule(cert_type=rule.cert_type, points=rule.points,
                             updated_by=admin.id, updated_at=datetime.utcnow()).insert()
    return {"message": f"{len(body.rules)} credit rules upserted"}


@router.delete("/credit-rules/{rule_id}")
async def delete_credit_rule(rule_id: PydanticObjectId, _user: User = _admin):
    rule = await CreditRule.get(rule_id)
    if not rule:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Credit rule not found")

    await rule.delete()
    return {"message": f"Credit rule '{rule.cert_type}' deleted"}


@router.post("/credits/reset")
async def reset_credit_points(body: CreditResetRequest, admin: User = _admin):
    semester = body.semester.strip()
    if not semester:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Semester is required")

    if not verify_password(body.admin_password, admin.password_hash):
        raise HTTPException(status.HTTP_401_UNAUTHORIZED, "Invalid super admin password")

    student_credit_docs = await StudentCredit.find_all().to_list()
    student_docs_updated = 0
    credit_entries_updated = 0
    for doc in student_credit_docs:
        updated_history = []
        history_changed = False
        for entry in doc.credit_history or []:
            if int(entry.points_awarded or 0) != 0:
                history_changed = True
                credit_entries_updated += 1
                updated_history.append(entry.model_copy(update={"points_awarded": 0}))
            else:
                updated_history.append(entry)

        if int(doc.total_credits or 0) != 0 or history_changed:
            await doc.set({
                "total_credits": 0,
                "credit_history": updated_history,
                "last_updated": datetime.utcnow(),
            })
            student_docs_updated += 1

    manual_submissions = await ManualCreditSubmission.find_all().to_list()
    manual_submissions_updated = 0
    for submission in manual_submissions:
        if int(submission.points_awarded or 0) == 0:
            continue
        await submission.set({"points_awarded": 0})
        manual_submissions_updated += 1

    return {
        "message": f"Credit points reset to 0 for semester {semester}.",
        "semester": semester,
        "student_documents_updated": student_docs_updated,
        "credit_entries_updated": credit_entries_updated,
        "manual_submissions_updated": manual_submissions_updated,
        "requested_by": admin.email,
    }


# ═══ PLATFORM STATS ══════════════════════════════════════════════════════════


@router.get("/stats")
async def platform_stats(_user: User = _admin):
    """Return aggregated platform-wide statistics for the admin overview dashboard."""
    from datetime import time as _time

    # Today 00:00:00 UTC — used as the lower bound for "today" counts
    today_start = datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0)

    # ── Simple counts ────────────────────────────────────────────────────
    total_clubs = await Club.find(Club.is_active == True).count()
    total_users = await User.find(User.role != UserRole.SUPER_ADMIN).count()
    total_students = await User.find(User.role == UserRole.STUDENT).count()
    club_certificates = await Certificate.find_all().count()
    dept_certificates = await DeptCertificate.find_all().count()
    guest_sessions = await GuestSession.find_all().to_list()
    guest_certificates = sum(len(session.guest_generated_certs or []) for session in guest_sessions)
    total_certificates = club_certificates + dept_certificates + guest_certificates

    club_certificates_today = await Certificate.find(
        {"issued_at": {"$gte": today_start}}
    ).count()

    dept_certificates_today = await DeptCertificate.find(
        {"created_at": {"$gte": today_start}}
    ).count()

    guest_sessions_today = await GuestSession.find(
        {"created_at": {"$gte": today_start}}
    ).to_list()
    guest_certificates_today = sum(
        len(session.guest_generated_certs or []) for session in guest_sessions_today
    )
    certificates_today = club_certificates_today + dept_certificates_today + guest_certificates_today

    emails_sent_today = await EmailLog.find(
        {"sent_at": {"$gte": today_start}, "status": EmailStatus.SENT.value}
    ).count()

    certs_by_source = [
        {"slug": "clubs", "name": "Clubs", "count": club_certificates},
        {"slug": "departments", "name": "Departments", "count": dept_certificates},
        {"slug": "guest", "name": "Guest", "count": guest_certificates},
    ]

    def _summarize_status(status_counts: dict[str, int], mailed_count: int, total_count: int) -> str:
        if total_count > 0 and mailed_count == total_count:
            return "emailed"
        if status_counts.get("generated", 0) > 0:
            return "generated"
        if status_counts.get("pending", 0) > 0:
            return "pending"
        if status_counts.get("failed", 0) > 0:
            return "failed"
        if status_counts.get("revoked", 0) > 0:
            return "revoked"
        return "generated"

    grouped_recent: dict[tuple[str, str, str], dict] = {}

    recent_club_certs = await Certificate.find_all().sort("-issued_at").limit(1200).to_list()
    for cert in recent_club_certs:
        generated_at = cert.issued_at
        if not generated_at:
            continue
        source_name = (cert.snapshot.club_name or "Club").strip() or "Club"
        event_name = (cert.snapshot.event_name or "Club Event").strip() or "Club Event"
        key = ("club", source_name, event_name)
        bucket = grouped_recent.get(key)
        if not bucket:
            bucket = {
                "source_type": "club",
                "source_name": source_name,
                "event_name": event_name,
                "generated_at": generated_at,
                "count": 0,
                "mailed_count": 0,
                "status_counts": {},
            }
            grouped_recent[key] = bucket
        if generated_at > bucket["generated_at"]:
            bucket["generated_at"] = generated_at
        bucket["count"] += 1
        status_key = cert.status.value
        bucket["status_counts"][status_key] = bucket["status_counts"].get(status_key, 0) + 1
        if status_key == CertStatus.EMAILED.value:
            bucket["mailed_count"] += 1

    recent_dept_certs = await DeptCertificate.find_all().sort("-created_at").limit(1200).to_list()
    dept_event_ids: list[PydanticObjectId] = []
    for cert in recent_dept_certs:
        if cert.event_id:
            try:
                dept_event_ids.append(PydanticObjectId(cert.event_id))
            except Exception:
                continue

    dept_event_map: dict[str, str] = {}
    if dept_event_ids:
        events = await DeptEvent.find({"_id": {"$in": dept_event_ids}}).to_list()
        dept_event_map = {str(evt.id): evt.name for evt in events}

    for cert in recent_dept_certs:
        generated_at = cert.created_at
        source_name = (cert.department or "Department").strip() or "Department"
        event_name = dept_event_map.get(str(cert.event_id), "Department Event") if cert.event_id else "Department Event"
        key = ("department", source_name, event_name)
        bucket = grouped_recent.get(key)
        if not bucket:
            bucket = {
                "source_type": "department",
                "source_name": source_name,
                "event_name": event_name,
                "generated_at": generated_at,
                "count": 0,
                "mailed_count": 0,
                "status_counts": {},
            }
            grouped_recent[key] = bucket
        if generated_at > bucket["generated_at"]:
            bucket["generated_at"] = generated_at
        bucket["count"] += 1
        status_key = "emailed" if cert.emailed_at else "generated"
        bucket["status_counts"][status_key] = bucket["status_counts"].get(status_key, 0) + 1
        if cert.emailed_at:
            bucket["mailed_count"] += 1

    recent_guest_sessions = await GuestSession.find_all().sort("-created_at").limit(600).to_list()
    for session in recent_guest_sessions:
        generated_count = len(session.guest_generated_certs or [])
        if generated_count <= 0:
            continue
        source_name = "Guest"
        event_name = (session.event_name or "Guest Event").strip() or "Guest Event"
        key = ("guest", source_name, event_name)
        bucket = grouped_recent.get(key)
        if not bucket:
            bucket = {
                "source_type": "guest",
                "source_name": source_name,
                "event_name": event_name,
                "generated_at": session.created_at,
                "count": 0,
                "mailed_count": 0,
                "status_counts": {},
            }
            grouped_recent[key] = bucket
        if session.created_at > bucket["generated_at"]:
            bucket["generated_at"] = session.created_at
        bucket["count"] += generated_count
        mailed_for_session = generated_count if session.guest_emails_sent else 0
        bucket["mailed_count"] += mailed_for_session
        status_key = "emailed" if session.guest_emails_sent else "generated"
        bucket["status_counts"][status_key] = bucket["status_counts"].get(status_key, 0) + generated_count

    recent_certificates = []
    for bucket in grouped_recent.values():
        total = int(bucket["count"] or 0)
        mailed = int(bucket["mailed_count"] or 0)
        recent_certificates.append({
            "source_type": bucket["source_type"],
            "source_name": bucket["source_name"],
            "event_name": bucket["event_name"],
            "generated_at": bucket["generated_at"],
            "count": total,
            "mailed_count": mailed,
            "status": _summarize_status(bucket["status_counts"], mailed, total),
        })

    recent_certificates.sort(key=lambda x: x["generated_at"] or datetime.min, reverse=True)
    recent_certificates = recent_certificates[:12]

    return {
        "total_clubs": total_clubs,
        "total_users": total_users,
        "total_students": total_students,
        "total_certificates": total_certificates,
        "certificates_today": certificates_today,
        "emails_sent_today": emails_sent_today,
        "certs_by_source": certs_by_source,
        "recent_certificates": recent_certificates,
    }
