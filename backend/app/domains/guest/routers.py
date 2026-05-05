"""
Guest Flow Router — 5-step certificate wizard backend.

All routes are prefixed: /guest
All routes are protected by require_guest.

POST /guest/start-session    — Create a new GuestSession (event name entry)
POST /guest/template         — Step 1: Upload custom PNG background template
POST /guest/excel            — Step 2a: Upload XLSX, parse headers + data
POST /guest/config           — Step 2b: Save selected_columns + email_column
POST /guest/field-positions  — Step 3: Save field positions for this session
POST /guest/generate         — Step 4: Pillow batch generation
GET  /guest/zip              — Step 5: Download all generated PNGs as ZIP (current session)
POST /guest/send-emails      — Step 5: Email each certificate to the email column
GET  /guest/status           — Poll current guest session state

GET  /guest/history                      — List all non-expired sessions for this guest
GET  /guest/sessions/{session_id}/zip    — Re-download ZIP for a past session
"""

import asyncio
import io
import logging
import re
import uuid
import zipfile
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, Dict, List, Optional

from beanie import PydanticObjectId
from fastapi import APIRouter, Body, Depends, File, HTTPException, UploadFile, status
from fastapi.responses import FileResponse, StreamingResponse
from pydantic import BaseModel, field_validator

from ...config import get_settings
from ...core.dependencies import require_guest
from ...models.dept_certificate import DeptCertificate
from ...models.field_position import FieldPosition
from ...models.guest_session import GuestSession
from ...models.student_credit import CreditHistoryEntry, StudentCredit
from ...models.user import User, UserRole
from ...services.storage_service import storage_path_to_url
from ...services.semester_service import get_current_semester
from ...services.email_service import send_certificate_email

logger = logging.getLogger(__name__)
settings = get_settings()

router = APIRouter(
    prefix="/guest",
    tags=["Guest Flow"],
)


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

def _guest_certs_dir(session_id: str) -> Path:
    """Directory where generated guest certificates are stored for a session."""
    d = settings.certs_dir / "guest" / session_id
    d.mkdir(parents=True, exist_ok=True)
    return d


def _guest_templates_dir() -> Path:
    """Directory where guest-uploaded PNG templates are stored."""
    d = settings.storage_root / "guest_templates"
    d.mkdir(parents=True, exist_ok=True)
    return d


async def _resolve_active_session(user: User) -> GuestSession:
    """Return the most recent non-expired GuestSession for this user.

    Raises HTTP 400 if no active session exists — the guest must call
    POST /guest/start-session first.
    """
    session = await GuestSession.find_one(
        GuestSession.user_id == user.id,
        GuestSession.expires_at > datetime.utcnow(),
        sort=[("created_at", -1)],
    )
    if not session:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            "No active guest session. Please start a new session from the dashboard.",
        )
    return session


class SendGuestEmailsRequest(BaseModel):
    row_indexes: Optional[List[int]] = None


class GuestGenerateRequest(BaseModel):
    allocate_points: bool = False
    points_per_cert: int = 0


_EMAIL_HEADER_PRIORITY = (
    "email",
    "e-mail",
    "mail",
    "email address",
    "student email",
    "participant email",
    "recipient email",
    "contact email",
)

_EMAIL_RE = re.compile(r"^[^\s@]+@[^\s@]+\.[^\s@]+$")


def _normalize_header(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().lower())


def _detect_email_column(rows: list[dict[str, Any]] | None, headers: list[str] | None = None) -> Optional[str]:
    header_list = headers or []
    normalized_headers = {header: _normalize_header(header) for header in header_list}

    for header, normalized in normalized_headers.items():
        if normalized in _EMAIL_HEADER_PRIORITY:
            return header

    for header, normalized in normalized_headers.items():
        if "email" in normalized or "e-mail" in normalized:
            return header

    if rows and header_list:
        best_header = None
        best_hits = 0
        for header in header_list:
            hits = 0
            for row in rows:
                value = str(row.get(header, "") or "").strip()
                if value and _EMAIL_RE.match(value):
                    hits += 1
            if hits > best_hits:
                best_header = header
                best_hits = hits

        if best_header and best_hits > 0:
            return best_header

    return None


def _safe_session_cert_path(session: GuestSession, cert_path_str: str) -> Path:
    cert_path = Path(cert_path_str)
    base_dir = _guest_certs_dir(str(session.id)).resolve()
    resolved = cert_path.resolve()
    if base_dir not in resolved.parents and resolved != base_dir:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Invalid certificate file path")
    return resolved


def _guest_session_cert_preview_url(session_id: str, file_name: str) -> str:
    return f"/guest/sessions/{session_id}/certificates/{file_name}"


# ─────────────────────────────────────────────────────────────────────────────
# START SESSION
# POST /guest/start-session
# ─────────────────────────────────────────────────────────────────────────────

class StartSessionRequest(BaseModel):
    event_name: str

    @field_validator("event_name")
    @classmethod
    def validate_event_name(cls, v: str) -> str:
        v = v.strip()
        if len(v) < 3:
            raise ValueError("Event name must be at least 3 characters")
        if len(v) > 100:
            raise ValueError("Event name must be at most 100 characters")
        return v


@router.post("/start-session")
async def start_guest_session(
    body: StartSessionRequest,
    current_user: User = Depends(require_guest),
):
    """Create a new GuestSession for this guest user.

    Each call creates a fresh session; previous sessions remain accessible
    in history until they expire (15 days).
    """
    now = datetime.utcnow()
    session = GuestSession(
        user_id=current_user.id,
        event_name=body.event_name,
        created_at=now,
        expires_at=now + timedelta(days=15),
    )
    await session.insert()
    return {
        "session_id": str(session.id),
        "event_name": session.event_name,
        "expires_at": session.expires_at.isoformat(),
    }


# ─────────────────────────────────────────────────────────────────────────────
# STEP 1 — Upload custom PNG template
# POST /guest/template
# ─────────────────────────────────────────────────────────────────────────────

@router.post("/template")
async def upload_guest_template(
    file: UploadFile = File(...),
    current_user: User = Depends(require_guest),
):
    """Upload a PNG file as the guest's custom certificate background."""
    session = await _resolve_active_session(current_user)

    if not file.content_type or not file.content_type.startswith("image/"):
        raise HTTPException(
            status.HTTP_422_UNPROCESSABLE_ENTITY,
            "Only image files (PNG/JPG) are accepted as templates",
        )

    content = await file.read()
    if len(content) > 1 * 1024 * 1024:
        raise HTTPException(
            status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
            "Template image must be smaller than 1 MB",
        )

    ext = Path(file.filename or "template.png").suffix or ".png"
    filename = f"guest_{session.id}_{uuid.uuid4().hex[:8]}{ext}"
    dest = _guest_templates_dir() / filename
    dest.write_bytes(content)

    await session.set({"guest_template_path": str(dest)})

    preview_url = f"/storage/guest_templates/{filename}"
    return {
        "message": "Template uploaded",
        "guest_template_path": str(dest),
        "preview_url": preview_url,
    }


# ─────────────────────────────────────────────────────────────────────────────
# STEP 2a — Upload XLSX, parse and return headers
# POST /guest/excel
# ─────────────────────────────────────────────────────────────────────────────

@router.post("/excel")
async def upload_guest_excel(
    file: UploadFile = File(...),
    current_user: User = Depends(require_guest),
):
    """Parse an uploaded XLSX file and return its column headers + row count."""
    session = await _resolve_active_session(current_user)

    content = await file.read()
    if len(content) > 20 * 1024 * 1024:
        raise HTTPException(
            status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
            "Excel file must be smaller than 20 MB",
        )

    try:
        from openpyxl import load_workbook
        wb = load_workbook(filename=io.BytesIO(content), read_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        headers_row = next(rows_iter, None)
        if not headers_row:
            raise HTTPException(status.HTTP_400_BAD_REQUEST, "Excel file is empty")

        headers = [str(h).strip() for h in headers_row if h is not None]
        if not headers:
            raise HTTPException(
                status.HTTP_400_BAD_REQUEST, "No column headers found in the first row"
            )

        parsed_rows: List[Dict] = []
        for row in rows_iter:
            if all(cell is None for cell in row):
                continue
            record: Dict = {}
            for col_idx, cell_value in enumerate(row):
                if col_idx < len(headers) and headers[col_idx]:
                    record[headers[col_idx]] = (
                        str(cell_value).strip() if cell_value is not None else ""
                    )
            parsed_rows.append(record)

        wb.close()
    except HTTPException:
        raise
    except Exception as exc:
        logger.error("Excel parse error: %s", exc)
        raise HTTPException(status.HTTP_400_BAD_REQUEST, f"Failed to parse Excel: {exc}")

    email_column = _detect_email_column(parsed_rows, headers)

    await session.set({
        "guest_excel_data": parsed_rows,
        "guest_selected_columns": None,
        "guest_email_column": email_column,
        "guest_generated_certs": None,
        "guest_emails_sent": False,
        "guest_email_statuses": None,
        "guest_allocate_points": False,
        "guest_points_per_cert": 0,
    })

    return {
        "headers": headers,
        "row_count": len(parsed_rows),
        "email_column": email_column,
        "message": f"Parsed {len(parsed_rows)} data rows",
    }


# ─────────────────────────────────────────────────────────────────────────────
# STEP 2b — Save column selection + email column
# POST /guest/config
# ─────────────────────────────────────────────────────────────────────────────

class GuestConfigRequest(BaseModel):
    selected_columns: List[str]
    email_column: Optional[str] = None


@router.post("/config")
async def save_guest_config(
    body: GuestConfigRequest,
    current_user: User = Depends(require_guest),
):
    """Persist which columns should be printed and which is the email column."""
    session = await _resolve_active_session(current_user)

    selected_columns = [c.strip() for c in (body.selected_columns or []) if isinstance(c, str) and c.strip()]
    if not selected_columns:
        raise HTTPException(
            status.HTTP_422_UNPROCESSABLE_ENTITY,
            "At least one column must be selected",
        )

    email_column = (body.email_column or "").strip() or None
    if not email_column and session.guest_excel_data:
        email_column = _detect_email_column(session.guest_excel_data, selected_columns) or session.guest_email_column

    await session.set({
        "guest_selected_columns": selected_columns,
        "guest_email_column": email_column,
        "guest_generated_certs": None,
        "guest_emails_sent": False,
        "guest_email_statuses": None,
        "guest_allocate_points": False,
        "guest_points_per_cert": 0,
    })

    return {
        "message": "Column configuration saved",
        "selected_columns": selected_columns,
        "email_column": email_column,
    }


# ─────────────────────────────────────────────────────────────────────────────
# STEP 3 — Save field positions for this session
# POST /guest/field-positions
# ─────────────────────────────────────────────────────────────────────────────

class ColumnPositionEntry(BaseModel):
    x_percent: float
    y_percent: float
    font_size: float = 24.0


class GuestFieldPositionsRequest(BaseModel):
    column_positions: Dict[str, ColumnPositionEntry]
    display_width: float = 580.0
    confirmed: bool = True


@router.post("/field-positions")
async def save_guest_field_positions(
    body: GuestFieldPositionsRequest,
    current_user: User = Depends(require_guest),
):
    """Save or update field positions for the current guest session.

    Uses session.id as the event_id in FieldPosition, cert_type='guest'.
    """
    session = await _resolve_active_session(current_user)

    # Serialise Pydantic models to plain dicts
    column_positions_dict: Dict[str, Dict[str, float]] = {
        col: {
            "x_percent": pos.x_percent,
            "y_percent": pos.y_percent,
            "font_size": pos.font_size,
        }
        for col, pos in body.column_positions.items()
    }

    # Upsert: update if exists, insert if not
    existing = await FieldPosition.find_one(
        FieldPosition.event_id == session.id,
        FieldPosition.cert_type == "guest",
    )
    if existing:
        await existing.set({
            "column_positions": column_positions_dict,
            "display_width": body.display_width,
            "confirmed": body.confirmed,
        })
    else:
        await FieldPosition(
            event_id=session.id,
            cert_type="guest",
            template_filename="__guest__",
            column_positions=column_positions_dict,
            display_width=body.display_width,
            confirmed=body.confirmed,
        ).insert()

    return {"message": "Field positions saved", "session_id": str(session.id)}


# ─────────────────────────────────────────────────────────────────────────────
# STEP 3.5 — Generate sample certificate from first row
# POST /guest/sample-preview
# ─────────────────────────────────────────────────────────────────────────────

@router.post("/sample-preview")
async def generate_guest_sample_preview(
    current_user: User = Depends(require_guest),
):
    """Generate and return a preview certificate using the first Excel row.

    This allows the guest to review placement/font sizing before batch generation.
    """
    session = await _resolve_active_session(current_user)

    if not session.guest_template_path or not Path(session.guest_template_path).exists():
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            "No valid certificate template found. Please complete Step 1 first.",
        )
    if not session.guest_excel_data:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            "No Excel data found. Please complete Step 2 first.",
        )
    if not session.guest_selected_columns:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            "Column configuration not saved. Please complete Step 2 first.",
        )

    fp = await FieldPosition.find_one(
        FieldPosition.event_id == session.id,
        FieldPosition.cert_type == "guest",
    )
    if not fp:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            "Field positions not configured. Please complete Step 3 first.",
        )

    first_row = session.guest_excel_data[0] if session.guest_excel_data else None
    if not first_row:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Excel does not contain data rows")

    sample_dir = settings.storage_root / "tmp"
    sample_dir.mkdir(parents=True, exist_ok=True)
    sample_name = f"guest_sample_{session.id}.png"
    sample_path = sample_dir / sample_name

    await asyncio.to_thread(
        _render_guest_certificate,
        Path(session.guest_template_path),
        first_row,
        fp.column_positions,
        str(sample_path),
    )

    first_row_selected = {
        col: str(first_row.get(col, "") if first_row.get(col) is not None else "")
        for col in (session.guest_selected_columns or [])
    }

    return {
        "sample_url": f"/storage/tmp/{sample_name}",
        "first_row": first_row_selected,
        "message": "Sample certificate generated from first row",
    }


# ─────────────────────────────────────────────────────────────────────────────
# STEP 4 — Generate certificates (Pillow batch)
# POST /guest/generate
# ─────────────────────────────────────────────────────────────────────────────

@router.post("/generate")
async def generate_guest_certificates(
    payload: GuestGenerateRequest = Body(default=GuestGenerateRequest()),
    current_user: User = Depends(require_guest),
):
    """Generate one PNG certificate per Excel row using the Pillow pipeline."""
    session = await _resolve_active_session(current_user)

    if not session.guest_template_path or not Path(session.guest_template_path).exists():
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            "No valid certificate template found. Please complete Step 1 first.",
        )
    if not session.guest_excel_data:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            "No Excel data found. Please complete Step 2 first.",
        )
    if not session.guest_selected_columns:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            "Column configuration not saved. Please complete Step 2 first.",
        )

    fp = await FieldPosition.find_one(
        FieldPosition.event_id == session.id,
        FieldPosition.cert_type == "guest",
    )
    if not fp:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            "Field positions not configured. Please complete Step 3 first.",
        )

    template_path = Path(session.guest_template_path)
    output_dir = _guest_certs_dir(str(session.id))
    rows = session.guest_excel_data

    generated: List[str] = []
    errors: List[str] = []
    credits_awarded = 0

    allocate_points = bool(payload.allocate_points)
    points_per_cert = int(payload.points_per_cert or 0)
    if points_per_cert < 0:
        points_per_cert = 0

    email_col = session.guest_email_column or _detect_email_column(rows, session.guest_selected_columns)
    if email_col and email_col != session.guest_email_column:
        await session.set({"guest_email_column": email_col})

    if allocate_points and not email_col:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            "Email column is required to allocate credit points.",
        )
    if allocate_points and points_per_cert <= 0:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            "Points per certificate must be greater than 0 when credit allocation is enabled.",
        )

    department_label = f"Guest Event - {session.event_name}" if session.event_name else "Guest Event"

    for idx, row in enumerate(rows):
        safe_name = f"cert_{idx+1:04d}_{uuid.uuid4().hex[:6]}.png"
        out_path = output_dir / safe_name
        try:
            await asyncio.to_thread(
                _render_guest_certificate,
                template_path,
                row,
                fp.column_positions,
                str(out_path),
            )
            generated.append(str(out_path))

            cert_number = f"GUEST-{str(session.id)[-6:]}-{idx + 1:04d}"
            participant_email = _normalize_guest_email(row.get(email_col) if email_col else "")
            student_name = _pick_row_value(row, ("Name", "name", "Full Name", "full_name", "Student Name")) or "Unknown"
            class_name = _pick_row_value(row, ("Class", "Department", "Semester", "Year", "Section")) or "-"
            contribution = _pick_row_value(row, ("Contribution", "Role", "Participation", "Certificate Type")) or "Participant"

            png_url = storage_path_to_url(str(out_path))
            existing = await DeptCertificate.find_one({"cert_number": cert_number})
            if existing:
                await existing.set({
                    "department": department_label,
                    "coordinator_user_id": str(current_user.id),
                    "event_id": str(session.id),
                    "name": student_name,
                    "class_name": class_name,
                    "contribution": contribution,
                    "participant_email": participant_email or None,
                    "png_url": png_url,
                })
            else:
                await DeptCertificate(
                    cert_number=cert_number,
                    department=department_label,
                    coordinator_user_id=str(current_user.id),
                    event_id=str(session.id),
                    name=student_name,
                    class_name=class_name,
                    contribution=contribution,
                    participant_email=participant_email or None,
                    png_url=png_url,
                    created_at=datetime.utcnow(),
                ).insert()

            # Credits removed from generation step - now awarded after email sending
        except Exception as exc:
            logger.error("Error generating cert for row %d: %s", idx + 1, exc)
            errors.append(f"Row {idx + 1}: {exc}")

    await session.set({
        "guest_generated_certs": generated,
        "guest_emails_sent": False,
        "guest_email_statuses": None,
        "guest_allocate_points": allocate_points,
        "guest_points_per_cert": points_per_cert,
    })

    return {
        "generated": len(generated),
        "errors": errors,
        "credits_awarded": credits_awarded,
        "message": f"Generated {len(generated)} certificate(s).",
    }


def _render_guest_certificate(
    template_path: Path,
    fields: dict,
    column_positions: dict,
    output_path: str,
) -> None:
    """Synchronous Pillow rendering — called inside asyncio.to_thread."""
    from PIL import Image, ImageDraw, ImageFont

    _FONTS_DIR = Path(__file__).resolve().parents[2] / "static" / "fonts"
    _FONT_CANDIDATES = [
        _FONTS_DIR / "Montserrat-Bold.ttf",
        _FONTS_DIR / "PlayfairDisplay.ttf",
        _FONTS_DIR / "EBGaramond.ttf",
        _FONTS_DIR / "Roboto.ttf",
    ]
    DEFAULT_FONT_PERCENT = 2.7

    def _load_font(size: int):
        for font_path in _FONT_CANDIDATES:
            try:
                if font_path.exists():
                    return ImageFont.truetype(str(font_path), size)
            except Exception:
                continue
        return ImageFont.load_default()

    img = Image.open(str(template_path)).convert("RGBA")
    draw = ImageDraw.Draw(img)
    img_w, img_h = img.size

    for col_header, pos in (column_positions or {}).items():
        value = str((fields or {}).get(col_header, ""))
        if not value:
            continue
        raw_font_size = pos.get("font_size")
        if raw_font_size is None:
            fsp = float(pos.get("font_size_percent") or DEFAULT_FONT_PERCENT)
            fsp = max(1.0, min(fsp, 8.0))
            font_size = max(10, int(img_w * fsp / 100))
        else:
            font_size = max(8, int(float(raw_font_size)))
        font = _load_font(font_size)
        x = (pos["x_percent"] / 100) * img_w
        y = (pos["y_percent"] / 100) * img_h
        draw.text(
            (int(x), int(y)),
            value,
            font=font,
            fill=(30, 30, 30, 255),
            anchor="mm",
            stroke_width=1,
            stroke_fill=(30, 30, 30, 255),
        )

    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    if img.mode == "RGBA":
        bg = Image.new("RGB", img.size, "white")
        bg.paste(img, mask=img.split()[3])
        final = bg
    else:
        final = img.convert("RGB")
    final.save(str(out), format="PNG", optimize=False)


# ─────────────────────────────────────────────────────────────────────────────
# STEP 5a — Download all certificates as ZIP (current session)
# GET /guest/zip
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/zip")
async def download_guest_zip(
    current_user: User = Depends(require_guest),
):
    """Return all generated guest certificates as a single ZIP (current session)."""
    session = await _resolve_active_session(current_user)
    return _build_zip_response(session)


# ─────────────────────────────────────────────────────────────────────────────
# STEP 5b — Send emails
# POST /guest/send-emails
# ─────────────────────────────────────────────────────────────────────────────

@router.post("/send-emails")
async def send_guest_emails(
    payload: Optional[SendGuestEmailsRequest] = Body(default=None),
    current_user: User = Depends(require_guest),
):
    """Send each generated certificate to the email address in the email column."""
    session = await _resolve_active_session(current_user)
    return await _send_guest_emails_for_session(session, payload)


async def _send_guest_emails_for_session(
    session: GuestSession,
    payload: Optional[SendGuestEmailsRequest] = None,
) -> Dict[str, Any]:
    """Send guest certificate emails for one session and persist row status."""

    if not session.guest_generated_certs:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            "No certificates available. Please complete Step 4 first.",
        )

    rows = session.guest_excel_data or []
    certs = session.guest_generated_certs or []
    email_col = session.guest_email_column or _detect_email_column(rows, session.guest_selected_columns)

    if email_col and email_col != session.guest_email_column:
        await session.set({"guest_email_column": email_col})

    if not email_col:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            "Email column not detected. Ensure the uploaded Excel contains an email column.",
        )

    if rows:
        header_keys = set(rows[0].keys())
        if email_col not in header_keys:
            raise HTTPException(
                status.HTTP_400_BAD_REQUEST,
                f"Configured email column '{email_col}' was not found in uploaded Excel headers.",
            )

    def _build_statuses() -> List[Dict[str, Any]]:
        statuses: List[Dict[str, Any]] = []
        for idx, row in enumerate(rows):
            recipient_email = (row.get(email_col, "") if email_col else "").strip()
            recipient_name = ""
            for name_key in ("Name", "name", "Full Name", "full_name", "Student Name"):
                if name_key in row and row[name_key]:
                    recipient_name = row[name_key]
                    break
            if not recipient_name and recipient_email:
                recipient_name = recipient_email.split("@")[0]
            statuses.append(
                {
                    "row_index": idx,
                    "recipient_email": recipient_email,
                    "recipient_name": recipient_name,
                    "status": "pending",
                    "sent_at": None,
                    "error": None,
                }
            )
        return statuses

    email_statuses = session.guest_email_statuses
    if not isinstance(email_statuses, list) or len(email_statuses) != len(rows):
        email_statuses = _build_statuses()
    else:
        for idx, row in enumerate(rows):
            entry = email_statuses[idx] if isinstance(email_statuses[idx], dict) else {}
            recipient_email = (row.get(email_col, "") if email_col else "").strip()
            recipient_name = ""
            for name_key in ("Name", "name", "Full Name", "full_name", "Student Name"):
                if name_key in row and row[name_key]:
                    recipient_name = row[name_key]
                    break
            if not recipient_name and recipient_email:
                recipient_name = recipient_email.split("@")[0]
            entry.setdefault("row_index", idx)
            entry["recipient_email"] = recipient_email
            entry["recipient_name"] = recipient_name
            entry["status"] = (entry.get("status") or "pending").replace("sent", "emailed")
            entry.setdefault("sent_at", None)
            entry.setdefault("error", None)
            email_statuses[idx] = entry

    target_indexes = None
    if payload and payload.row_indexes:
        target_indexes = {idx for idx in payload.row_indexes if isinstance(idx, int)}

    sent = 0
    failed = 0
    errors: List[str] = []

    for idx, row in enumerate(rows):
        if target_indexes is not None and idx not in target_indexes:
            continue

        entry = email_statuses[idx]
        if entry.get("status") == "emailed":
            continue

        recipient_email = (entry.get("recipient_email") or "").strip()
        if not recipient_email:
            entry["status"] = "failed"
            entry["error"] = f"Row {idx + 1}: No email address found in column '{email_col}'"
            failed += 1
            errors.append(entry["error"])
            continue

        cert_path = certs[idx] if idx < len(certs) else None
        if not cert_path or not Path(cert_path).exists():
            entry["status"] = "failed"
            entry["error"] = f"Row {idx + 1}: Certificate file not found — {cert_path}"
            failed += 1
            errors.append(entry["error"])
            continue

        recipient_name = entry.get("recipient_name") or recipient_email.split("@")[0]
        cert_number = f"GUEST-{str(session.id)[-6:]}-{idx + 1:04d}"
        try:
            success = await send_certificate_email(
                recipient_email=recipient_email,
                recipient_name=recipient_name,
                cert_number=cert_number,
                event_name=session.event_name,
                club_name="Guest Event",
                png_path=cert_path,
            )
                if success:
                    entry["status"] = "emailed"
                    entry["sent_at"] = datetime.utcnow()
                    entry["error"] = None
                    
                    # Award credits after successful email
                    if session.guest_allocate_points and session.guest_points_per_cert > 0:
                        try:
                            await _award_guest_credits(
                                cert_number=cert_number,
                                student_email=recipient_email,
                                student_name=recipient_name,
                                points=session.guest_points_per_cert,
                                event_name=session.event_name,
                            )
                        except Exception as exc:
                            logger.error("Failed to award credits for %s after email: %s", recipient_email, exc)

                    existing = await DeptCertificate.find_one({"cert_number": cert_number})
                if existing:
                    await existing.set({
                        "emailed_at": entry["sent_at"],
                        "email_error": None,
                    })
                sent += 1
            else:
                entry["status"] = "failed"
                entry["error"] = f"Row {idx + 1}: Email delivery failed (rate limit or SMTP error)"
                existing = await DeptCertificate.find_one({"cert_number": cert_number})
                if existing:
                    await existing.set({"email_error": entry["error"]})
                failed += 1
                errors.append(entry["error"])
        except Exception as exc:
            entry["status"] = "failed"
            entry["error"] = f"Row {idx + 1}: {exc}"
            existing = await DeptCertificate.find_one({"cert_number": cert_number})
            if existing:
                await existing.set({"email_error": entry["error"]})
            failed += 1
            errors.append(entry["error"])

    sent_count = sum(1 for entry in email_statuses if entry.get("status") == "emailed")
    failed_count = sum(1 for entry in email_statuses if entry.get("status") == "failed")
    pending_count = sum(1 for entry in email_statuses if entry.get("status") not in ("emailed", "failed"))

    await session.set(
        {
            "guest_email_statuses": email_statuses,
            "guest_emails_sent": sent_count > 0,
        }
    )

    return {
        "sent": sent,
        "failed": failed,
        "errors": errors,
        "message": f"Sent {sent} email(s). {failed} failed.",
        "email_statuses": email_statuses,
        "email_sent_count": sent_count,
        "email_failed_count": failed_count,
        "email_pending_count": pending_count,
        "email_total_count": len(email_statuses),
    }


@router.post("/sessions/{session_id}/send-emails")
async def send_guest_session_emails(
    session_id: PydanticObjectId,
    payload: Optional[SendGuestEmailsRequest] = Body(default=None),
    current_user: User = Depends(require_guest),
):
    """Send unsent certificate emails for a historical guest session."""
    session = await GuestSession.get(session_id)
    if not session:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Session not found")
    if session.user_id != current_user.id:
        raise HTTPException(status.HTTP_403_FORBIDDEN, "Access denied")
    if session.expires_at < datetime.utcnow():
        raise HTTPException(status.HTTP_410_GONE, "Session has expired and files have been deleted")

    return await _send_guest_emails_for_session(session, payload)


# ─────────────────────────────────────────────────────────────────────────────
# STATUS — Poll current guest session state
# GET /guest/status
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/status")
async def get_guest_status(
    current_user: User = Depends(require_guest),
):
    """Return the current state of the active guest session."""
    session = await _resolve_active_session(current_user)

    fp = await FieldPosition.find_one(
        FieldPosition.event_id == session.id,
        FieldPosition.cert_type == "guest",
    )

    template_url = None
    if session.guest_template_path:
        tmpl = Path(session.guest_template_path)
        if tmpl.exists():
            template_url = f"/storage/guest_templates/{tmpl.name}"

    email_statuses = session.guest_email_statuses or []
    for entry in email_statuses:
        if isinstance(entry, dict) and entry.get("status") == "sent":
            entry["status"] = "emailed"
    email_column = session.guest_email_column or _detect_email_column(session.guest_excel_data, session.guest_selected_columns)
    if not email_statuses and session.guest_excel_data and email_column:
        rows = session.guest_excel_data
        email_col = email_column
        for idx, row in enumerate(rows):
            recipient_email = (row.get(email_col, "") if email_col else "").strip()
            recipient_name = ""
            for name_key in ("Name", "name", "Full Name", "full_name", "Student Name"):
                if name_key in row and row[name_key]:
                    recipient_name = row[name_key]
                    break
            if not recipient_name and recipient_email:
                recipient_name = recipient_email.split("@")[0]
            email_statuses.append(
                {
                    "row_index": idx,
                    "recipient_email": recipient_email,
                    "recipient_name": recipient_name,
                    "status": "pending",
                    "sent_at": None,
                    "error": None,
                }
            )

    return {
        "session_id": str(session.id),
        "event_name": session.event_name,
        "expires_at": session.expires_at.isoformat(),
        "step1_complete": bool(session.guest_template_path and Path(session.guest_template_path).exists()),
        "template_url": template_url,
        "step2_complete": bool(
            session.guest_excel_data
            and session.guest_selected_columns
        ),
        "excel_row_count": len(session.guest_excel_data) if session.guest_excel_data else 0,
        "selected_columns": session.guest_selected_columns or [],
        "email_column": email_column,
        "all_excel_headers": (
            list(session.guest_excel_data[0].keys()) if session.guest_excel_data else []
        ),
        "step3_complete": bool(fp),
        "field_positions": {
            "column_positions": fp.column_positions if fp else {},
            "display_width": fp.display_width if fp else 580.0,
        } if fp else None,
        "step4_complete": bool(session.guest_generated_certs),
        "generated_count": len(session.guest_generated_certs) if session.guest_generated_certs else 0,
        "step5_emails_sent": session.guest_emails_sent,
        "guest_allocate_points": bool(session.guest_allocate_points),
        "guest_points_per_cert": int(session.guest_points_per_cert or 0),
        "email_statuses": email_statuses,
        "email_sent_count": sum(1 for e in email_statuses if e.get("status") == "emailed"),
        "email_failed_count": sum(1 for e in email_statuses if e.get("status") == "failed"),
        "email_pending_count": sum(1 for e in email_statuses if e.get("status") not in ("emailed", "failed")),
        "email_total_count": len(email_statuses),
    }


# ─────────────────────────────────────────────────────────────────────────────
# HISTORY — List all non-expired sessions
# GET /guest/history
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/history")
async def get_guest_history(
    current_user: User = Depends(require_guest),
):
    """Return all non-expired GuestSessions for this guest, newest first."""
    now = datetime.utcnow()
    sessions = await GuestSession.find(
        GuestSession.user_id == current_user.id,
        GuestSession.expires_at > now,
        sort=[("created_at", -1)],
    ).to_list()

    result = []
    for s in sessions:
        cert_count = len(s.guest_generated_certs) if s.guest_generated_certs else 0
        has_downloadable = (
            cert_count > 0
            and any(Path(p).exists() for p in (s.guest_generated_certs or []))
        )
        days_remaining = max(0, (s.expires_at - now).days)
        result.append({
            "session_id": str(s.id),
            "event_name": s.event_name,
            "created_at": s.created_at.isoformat(),
            "expires_at": s.expires_at.isoformat(),
            "days_remaining": days_remaining,
            "cert_count": cert_count,
            "emails_sent": s.guest_emails_sent,
            "has_downloadable_certs": has_downloadable,
        })

    return result


@router.get("/sessions/{session_id}")
async def get_guest_session_detail(
    session_id: PydanticObjectId,
    current_user: User = Depends(require_guest),
):
    """Return details for one historical guest session."""
    session = await GuestSession.get(session_id)
    if not session:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Session not found")
    if session.user_id != current_user.id:
        raise HTTPException(status.HTTP_403_FORBIDDEN, "Access denied")
    if session.expires_at < datetime.utcnow():
        raise HTTPException(status.HTTP_410_GONE, "Session has expired and files have been deleted")

    cert_paths = session.guest_generated_certs or []
    email_statuses = session.guest_email_statuses or []
    rows = session.guest_excel_data or []

    certificates = []
    for idx, cert_path_str in enumerate(cert_paths):
        cert_path = Path(cert_path_str)
        if not cert_path.exists():
            continue

        email_entry = email_statuses[idx] if idx < len(email_statuses) and isinstance(email_statuses[idx], dict) else {}
        row = rows[idx] if idx < len(rows) and isinstance(rows[idx], dict) else {}
        recipient_email = (email_entry.get("recipient_email") or "").strip() or next(
            (str(row.get(col, "") or "").strip() for col in (session.guest_email_column, ) if col and str(row.get(col, "") or "").strip()),
            "",
        )
        recipient_name = (email_entry.get("recipient_name") or "").strip() or next(
            (str(row.get(key, "") or "").strip() for key in ("Name", "name", "Full Name", "full_name", "Student Name") if str(row.get(key, "") or "").strip()),
            recipient_email.split("@")[0] if recipient_email else "",
        )
        certificates.append({
            "index": idx,
            "cert_number": f"GUEST-{str(session.id)[-6:]}-{idx + 1:04d}",
            "recipient_email": recipient_email,
            "recipient_name": recipient_name,
            "status": (email_entry.get("status") or ("emailed" if session.guest_emails_sent else "generated")).replace("sent", "emailed"),
            "sent_at": email_entry.get("sent_at"),
            "error": email_entry.get("error"),
            "file_name": cert_path.name,
            "preview_url": _guest_session_cert_preview_url(str(session.id), cert_path.name),
        })

    return {
        "session_id": str(session.id),
        "event_name": session.event_name,
        "created_at": session.created_at.isoformat(),
        "expires_at": session.expires_at.isoformat(),
        "days_remaining": max(0, (session.expires_at - datetime.utcnow()).days),
        "cert_count": len(certificates),
        "emails_sent": session.guest_emails_sent,
        "guest_allocate_points": bool(session.guest_allocate_points),
        "guest_points_per_cert": int(session.guest_points_per_cert or 0),
        "email_column": session.guest_email_column,
        "certificates": certificates,
    }


@router.get("/sessions/{session_id}/certificates/{file_name}")
async def download_guest_session_certificate(
    session_id: PydanticObjectId,
    file_name: str,
    current_user: User = Depends(require_guest),
):
    """Return a single generated certificate image for preview or download."""
    session = await GuestSession.get(session_id)
    if not session:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Session not found")
    if session.user_id != current_user.id:
        raise HTTPException(status.HTTP_403_FORBIDDEN, "Access denied")
    if session.expires_at < datetime.utcnow():
        raise HTTPException(status.HTTP_410_GONE, "Session has expired and files have been deleted")

    if not session.guest_generated_certs:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "No certificates found for this session")

    target = None
    for cert_path_str in session.guest_generated_certs:
        cert_path = _safe_session_cert_path(session, cert_path_str)
        if cert_path.name == file_name:
            target = cert_path
            break

    if not target or not target.exists():
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Certificate file not found")

    return FileResponse(
        path=str(target),
        media_type="image/png",
        filename=target.name,
    )


# ─────────────────────────────────────────────────────────────────────────────
# HISTORY ZIP — Re-download ZIP for a specific past session
# GET /guest/sessions/{session_id}/zip
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/sessions/{session_id}/zip")
async def download_session_zip(
    session_id: PydanticObjectId,
    current_user: User = Depends(require_guest),
):
    """Re-download ZIP for a historical session. Ownership-checked."""
    session = await GuestSession.get(session_id)
    if not session:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Session not found")
    if session.user_id != current_user.id:
        raise HTTPException(status.HTTP_403_FORBIDDEN, "Access denied")
    if session.expires_at < datetime.utcnow():
        raise HTTPException(status.HTTP_410_GONE, "Session has expired and files have been deleted")
    return _build_zip_response(session)


# ─────────────────────────────────────────────────────────────────────────────
# Internal: build StreamingResponse ZIP
# ─────────────────────────────────────────────────────────────────────────────

def _build_zip_response(session: GuestSession) -> StreamingResponse:
    """Build a ZIP StreamingResponse from a session's generated certs."""
    if not session.guest_generated_certs:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            "No certificates have been generated yet. Please complete Step 4 first.",
        )

    existing = [p for p in session.guest_generated_certs if Path(p).exists()]
    if not existing:
        raise HTTPException(
            status.HTTP_404_NOT_FOUND,
            "Certificate files not found on disk. They may have been deleted.",
        )

    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for path_str in existing:
            p = Path(path_str)
            zf.write(p, arcname=p.name)
    zip_buf.seek(0)

    safe_name = session.event_name.replace(" ", "_")[:50]
    return StreamingResponse(
        zip_buf,
        media_type="application/zip",
        headers={
            "Content-Disposition": f'attachment; filename="{safe_name}_certificates.zip"'
        },
    )
