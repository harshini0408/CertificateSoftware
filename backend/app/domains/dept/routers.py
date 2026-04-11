from io import BytesIO
from pathlib import Path
from typing import Optional, Any, Dict
import hashlib
import re
from datetime import datetime
from uuid import uuid4

from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, UploadFile, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from PIL import Image, ImageDraw, ImageFont
from pydantic import BaseModel

from ...core.dependencies import require_role
from ...models.user import User, UserRole
from ...models.student_credit import StudentCredit
from ...models.dept_asset import DeptAsset
from ...models.dept_certificate import DeptCertificate
from ...models.dept_certificate_preview import DeptCertificatePreview
from ...models.dept_event import DeptEvent, DeptEventStatus
from ...models.dept_template import DeptTemplate
from ...services.signature_service import process_signature, save_logo
from ...services.storage_service import save_cert_png
from ...services.storage_service import storage_path_to_url
from ...services.storage_service import storage_url_to_path
from ...services.email_service import send_certificate_email
from ...config import get_settings

router = APIRouter(tags=["Department"])


def _clean_text_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))

    txt = str(value).strip()
    if txt.endswith(" 00:00:00") and re.match(r"^\d{4}-\d{2}-\d{2} 00:00:00$", txt):
        return txt.split(" ")[0]
    if re.match(r"^-?\d+\.0$", txt):
        return txt[:-2]
    return txt


class DeptBatchDownloadRequest(BaseModel):
    cert_numbers: list[str]


class DeptEventCreateRequest(BaseModel):
    name: str
    event_date: Optional[datetime] = None
    semester: Optional[str] = ""


class DeptEventFieldMappingRequest(BaseModel):
    selected_fields: list[str]
    field_positions: Dict[str, Dict[str, float]]


def _event_response(evt: DeptEvent) -> dict:
    return {
        "id": str(evt.id),
        "name": evt.name,
        "event_date": evt.event_date,
        "semester": evt.semester,
        "status": evt.status.value,
        "participant_count": evt.participant_count,
        "cert_count": evt.cert_count,
        "excel_headers": evt.excel_headers,
        "source_rows_count": len(evt.excel_rows or []),
        "preview_row": evt.excel_preview_row,
        "preview_certificate_id": str(evt.preview_certificate_id) if evt.preview_certificate_id else None,
        "preview_approved": bool(evt.preview_approved),
        "created_at": evt.created_at,
    }


def _slugify(value: str) -> str:
    text = re.sub(r"[^a-zA-Z0-9]+", "-", value.strip().lower())
    return text.strip("-") or "dept"


def _normalize_department(department: Optional[str]) -> str:
    normalized = (department or "").strip()
    if not normalized:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            "Department is not configured for this account. Ask Super Admin to set your department.",
        )
    return normalized


def _pick_single_template() -> Path:
    templates_dir = Path(__file__).parent.parent.parent / "static" / "certificate_templates"
    pngs = sorted(templates_dir.glob("*.png"))
    if not pngs:
        raise HTTPException(status.HTTP_500_INTERNAL_SERVER_ERROR, "No PNG certificate templates available")
    return pngs[0]


def _parse_dept_excel(file_bytes: bytes) -> list[dict]:
    wb = load_workbook(filename=BytesIO(file_bytes), read_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    headers_row = next(rows, None)
    if not headers_row:
        wb.close()
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Excel is empty")

    headers = [str(h).strip() if h is not None else "" for h in headers_row]
    norm = [h.lower().replace(" ", "").replace("_", "") for h in headers]

    def idx_for(*names: str) -> int:
        for i, key in enumerate(norm):
            if key in names:
                return i
        return -1

    name_i = idx_for("name")
    class_i = idx_for("class", "classname", "departmentclass")
    contribution_i = idx_for("contribution", "whatcontribution", "role")

    if name_i < 0 or class_i < 0 or contribution_i < 0:
        wb.close()
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            "Excel must include columns: Name, Class, Contribution",
        )

    parsed = []
    for row in rows:
        if not row or all(v is None for v in row):
            continue

        name = str(row[name_i]).strip() if name_i < len(row) and row[name_i] is not None else ""
        class_name = str(row[class_i]).strip() if class_i < len(row) and row[class_i] is not None else ""
        contribution = str(row[contribution_i]).strip() if contribution_i < len(row) and row[contribution_i] is not None else ""

        if not name or not class_name or not contribution:
            continue

        parsed.append(
            {
                "name": name,
                "class_name": class_name,
                "contribution": contribution,
            }
        )

    wb.close()
    if not parsed:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "No valid rows found in Excel")
    return parsed


def _parse_excel_headers(file_bytes: bytes) -> list[str]:
    wb = load_workbook(filename=BytesIO(file_bytes), read_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    headers_row = next(rows, None)
    wb.close()
    if not headers_row:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Excel is empty")
    headers = [str(h).strip() if h is not None else "" for h in headers_row]
    headers = [h for h in headers if h]
    if not headers:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "No usable header columns found")
    return headers


def _parse_excel_rows_dynamic(file_bytes: bytes) -> list[dict[str, str]]:
    wb = load_workbook(filename=BytesIO(file_bytes), read_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    headers_row = next(rows, None)
    if not headers_row:
        wb.close()
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Excel is empty")

    headers = [str(h).strip() if h is not None else "" for h in headers_row]
    if not any(headers):
        wb.close()
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Excel header row is empty")

    parsed: list[dict[str, str]] = []
    for row in rows:
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue
        rec: dict[str, str] = {}
        for i, h in enumerate(headers):
            if not h:
                continue
            val = row[i] if i < len(row) else None
            rec[h] = _clean_text_value(val)
        parsed.append(rec)

    wb.close()
    if not parsed:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "No valid rows found in Excel")
    return parsed


def _pick_email_from_row(row: dict[str, str]) -> Optional[str]:
    candidate_keys = ["Email", "E-mail", "Mail", "Student Email", "Participant Email"]
    lower_map = {k.lower(): v for k, v in row.items()}

    for k in candidate_keys:
        v = row.get(k)
        if isinstance(v, str) and "@" in v:
            return v.strip()

    for k, v in lower_map.items():
        if "mail" in k and isinstance(v, str) and "@" in v:
            return v.strip()

    return None


def _load_font(size: int):
    font_path = Path(__file__).parent.parent.parent / "static" / "fonts" / "PlayfairDisplay.ttf"
    try:
        if font_path.exists():
            return ImageFont.truetype(str(font_path), size)
        else:
            import logging
            logging.getLogger(__name__).warning("Font missing: %s. Using PIL default.", font_path)
    except Exception as exc:
        import logging
        logging.getLogger(__name__).warning("Failed to load font %s: %s", font_path, exc)
    return ImageFont.load_default()


def _render_dept_certificate(
    template_path: Path,
    name: str,
    class_name: str,
    contribution: str,
    cert_number: str,
    logo_path: Optional[str],
    sig1_path: Optional[str],
    sig2_path: Optional[str],
    field_positions: Optional[dict] = None,
) -> bytes:
    img = Image.open(str(template_path)).convert("RGBA")
    draw = ImageDraw.Draw(img)
    w, h = img.size

    def get_pos(field: str, def_x: float, def_y: float, def_size: int):
        if not field_positions or field not in field_positions:
            return def_x, def_y, def_size
        pos = field_positions[field]
        return (
            pos.get("x_percent", def_x * 100) / 100,
            pos.get("y_percent", def_y * 100) / 100,
            pos.get("font_size", def_size),
        )

    # Resolve positions and sizes (pixel-based defaults).
    nx, ny, ns = get_pos("name", 0.5, 0.46, 56)
    cx, cy, cs = get_pos("class_name", 0.5, 0.56, 44)
    tx, ty, ts = get_pos("contribution", 0.5, 0.64, 44)
    # cert_number is usually fixed at top right but can be made dynamic if needed
    zx, zy, zs = get_pos("cert_number", 0.82, 0.05, 24)

    name_font = _load_font(max(1, int(ns)))
    class_font = _load_font(max(1, int(cs)))
    contrib_font = _load_font(max(1, int(ts)))
    cert_font = _load_font(max(1, int(zs)))

    draw.text((w * nx, h * ny), name, fill=(28, 35, 70, 255), font=name_font, anchor="mm")
    draw.text((w * cx, h * cy), f"Class: {class_name}", fill=(45, 45, 45, 255), font=class_font, anchor="mm")
    draw.text((w * tx, h * ty), f"Contribution: {contribution}", fill=(45, 45, 45, 255), font=contrib_font, anchor="mm")
    draw.text((w * zx, h * zy), cert_number, fill=(44, 61, 127, 255), font=cert_font, anchor="lm")

    def _paste_scaled(path: Optional[str], field_id: str, def_x: float, def_y: float, max_w_ratio: float, max_h_ratio: float):
        if not path:
            return
        p = Path(path)
        if not p.exists():
            return
        
        # Resolve position
        cx, cy, _ = get_pos(field_id, def_x, def_y, 0)

        try:
            asset = Image.open(str(p)).convert("RGBA")
            max_w = int(w * max_w_ratio)
            max_h = int(h * max_h_ratio)
            asset.thumbnail((max_w, max_h), Image.LANCZOS)
            x = int(w * cx - asset.width / 2)
            y = int(h * cy - asset.height / 2)
            img.paste(asset, (x, y), asset.split()[3])
        except Exception:
            return

    _paste_scaled(logo_path, "logo", 0.12, 0.12, 0.16, 0.16)
    _paste_scaled(sig1_path, "signature_primary", 0.18, 0.84, 0.2, 0.1)
    _paste_scaled(sig2_path, "signature_secondary", 0.78, 0.84, 0.2, 0.1)

    out = BytesIO()
    if img.mode == "RGBA":
        bg = Image.new("RGB", img.size, "white")
        bg.paste(img, mask=img.split()[3])
        bg.save(out, format="PNG")
    else:
        img.save(out, format="PNG")
    out.seek(0)
    return out.getvalue()


def _render_dept_certificate_dynamic(
    template_path: Path,
    row: dict[str, str],
    cert_number: str,
    selected_fields: list[str],
    field_positions: Dict[str, Dict[str, float]],
    logo_path: Optional[str],
    sig1_path: Optional[str],
    event_date: Optional[datetime] = None,
) -> bytes:
    img = Image.open(str(template_path)).convert("RGBA")
    draw = ImageDraw.Draw(img)
    w, h = img.size

    for field in selected_fields:
        pos = field_positions.get(field)
        if not pos:
            continue
        x = float(pos.get("x_percent", 50.0))
        y = float(pos.get("y_percent", 50.0))
        font_size = int(float(pos.get("font_size", 36)))
        font = _load_font(max(1, font_size))

        if field == "_date":
            dt = event_date or datetime.utcnow()
            value = dt.strftime("%d-%m-%Y")
        else:
            value = _clean_text_value(row.get(field))

        if value:
            draw.text((w * x / 100, h * y / 100), value, fill=(28, 35, 70, 255), font=font, anchor="mm")

    cert_pos = field_positions.get("_cert_number")
    if cert_pos:
        cert_font = _load_font(max(1, int(float(cert_pos.get("font_size", 24)))))
        draw.text(
            (w * float(cert_pos.get("x_percent", 82.0)) / 100, h * float(cert_pos.get("y_percent", 5.0)) / 100),
            cert_number,
            fill=(44, 61, 127, 255),
            font=cert_font,
            anchor="lm",
        )

    def _paste_scaled(path: Optional[str], field_id: str, def_x: float, def_y: float, max_w_ratio: float, max_h_ratio: float):
        if not path:
            return
        p = Path(path)
        if not p.exists():
            return
        pos = field_positions.get(field_id, {})
        cx = float(pos.get("x_percent", def_x * 100)) / 100
        cy = float(pos.get("y_percent", def_y * 100)) / 100

        try:
            asset = Image.open(str(p)).convert("RGBA")
            max_w = int(w * max_w_ratio)
            max_h = int(h * max_h_ratio)
            asset.thumbnail((max_w, max_h), Image.LANCZOS)
            x = int(w * cx - asset.width / 2)
            y = int(h * cy - asset.height / 2)
            img.paste(asset, (x, y), asset.split()[3])
        except Exception:
            return

    if field_positions.get("_logo"):
        _paste_scaled(logo_path, "_logo", 0.12, 0.12, 0.16, 0.16)
    if field_positions.get("_signature"):
        _paste_scaled(sig1_path, "_signature", 0.18, 0.84, 0.2, 0.1)

    out = BytesIO()
    if img.mode == "RGBA":
        bg = Image.new("RGB", img.size, "white")
        bg.paste(img, mask=img.split()[3])
        bg.save(out, format="PNG")
    else:
        img.save(out, format="PNG")
    out.seek(0)
    return out.getvalue()


async def _get_or_create_dept_asset(department: str) -> DeptAsset:
    asset = await DeptAsset.find_one(DeptAsset.department == department)
    if asset:
        return asset

    asset = DeptAsset(department=department)
    await asset.insert()
    return asset


@router.get("/dept/assets")
async def get_dept_assets(
    current_user: User = Depends(require_role(UserRole.DEPT_COORDINATOR)),
):
    department = _normalize_department(current_user.department)
    asset = await _get_or_create_dept_asset(department)

    logo_url = asset.logo_url
    if asset.logo_path and (not logo_url or not str(logo_url).startswith("/storage/")):
        logo_url = storage_path_to_url(asset.logo_path)

    signature_url = asset.signature1_url
    if asset.signature1_path and (not signature_url or not str(signature_url).startswith("/storage/")):
        signature_url = storage_path_to_url(asset.signature1_path)

    if logo_url != asset.logo_url or signature_url != asset.signature1_url:
        asset.logo_url = logo_url
        asset.signature1_url = signature_url
        asset.updated_at = datetime.utcnow()
        await asset.save()

    return {
        "department": department,
        "logo_url": logo_url,
        "logo_hash": asset.logo_hash,
        "signature_url": signature_url,
        "signature_hash": asset.signature1_hash,
        "has_logo": bool(asset.logo_path),
        "has_signature": bool(asset.signature1_path),
    }


@router.post("/dept/assets")
async def upsert_dept_assets(
    logo: UploadFile | None = File(None),
    signature: UploadFile | None = File(None),
    current_user: User = Depends(require_role(UserRole.DEPT_COORDINATOR)),
):
    if logo is None and signature is None:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Upload at least one asset (logo/signature)")

    department = _normalize_department(current_user.department)
    dept_slug = _slugify(department)
    asset = await _get_or_create_dept_asset(department)

    if logo is not None:
        logo_bytes = await logo.read()
        asset.logo_path = save_logo(logo_bytes, dept_slug)
        asset.logo_hash = hashlib.md5(logo_bytes).hexdigest()
        asset.logo_url = storage_path_to_url(asset.logo_path)

    if signature is not None:
        sig_bytes = await signature.read()
        asset.signature1_path = process_signature(sig_bytes, dept_slug)
        asset.signature1_hash = hashlib.md5(sig_bytes).hexdigest()
        asset.signature1_url = storage_path_to_url(asset.signature1_path)

    asset.updated_at = datetime.utcnow()
    await asset.save()

    return {
        "message": "Department assets updated",
        "department": department,
        "logo_url": asset.logo_url,
        "logo_hash": asset.logo_hash,
        "signature_url": asset.signature1_url,
        "signature_hash": asset.signature1_hash,
        "has_logo": bool(asset.logo_path),
        "has_signature": bool(asset.signature1_path),
    }


@router.get("/dept/dashboard")
async def get_dept_dashboard(
    current_user: User = Depends(require_role(UserRole.DEPT_COORDINATOR)),
):
    department = _normalize_department(current_user.department)
    user_id = str(current_user.id)

    events = await DeptEvent.find(
        DeptEvent.department == department,
        DeptEvent.created_by_user_id == user_id,
    ).sort(-DeptEvent.created_at).to_list()

    event_ids = [str(evt.id) for evt in events]
    certs = await DeptCertificate.find(
        DeptCertificate.department == department,
        {"event_id": {"$in": event_ids}},
    ).to_list() if event_ids else []

    participants = {f"{(c.name or '').strip().lower()}|{(c.class_name or '').strip().lower()}" for c in certs if c.name}

    recent_events = [_event_response(evt) for evt in events[:5]]

    return {
        "department": department,
        "stats": {
            "total_events": len(events),
            "total_certificates_issued": len(certs),
            "total_participants": len(participants),
        },
        "recent_events": recent_events,
    }


@router.get("/dept/events")
async def list_dept_events(
    current_user: User = Depends(require_role(UserRole.DEPT_COORDINATOR)),
):
    department = _normalize_department(current_user.department)
    user_id = str(current_user.id)
    events = await DeptEvent.find(
        DeptEvent.department == department,
        DeptEvent.created_by_user_id == user_id,
    ).sort(-DeptEvent.created_at).to_list()
    return [_event_response(evt) for evt in events]


@router.get("/dept/events/{event_id}")
async def get_dept_event(
    event_id: str,
    current_user: User = Depends(require_role(UserRole.DEPT_COORDINATOR)),
):
    department = _normalize_department(current_user.department)
    evt = await _get_dept_event_or_404(event_id, department, str(current_user.id))
    return _event_response(evt)


@router.post("/dept/events", status_code=201)
async def create_dept_event(
    body: DeptEventCreateRequest,
    current_user: User = Depends(require_role(UserRole.DEPT_COORDINATOR)),
):
    department = _normalize_department(current_user.department)
    name = (body.name or "").strip()
    semester = (body.semester or "").strip()
    if not name:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Event name is required")

    evt = DeptEvent(
        department=department,
        created_by_user_id=str(current_user.id),
        name=name,
        event_date=body.event_date,
        semester=semester,
        status=DeptEventStatus.DRAFT,
        participant_count=0,
        cert_count=0,
    )
    await evt.insert()
    return _event_response(evt)


async def _get_dept_event_or_404(event_id: str, department: str, user_id: str) -> DeptEvent:
    try:
        evt = await DeptEvent.get(event_id)
    except Exception:
        evt = None
    if not evt or evt.department != department or evt.created_by_user_id != user_id:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Department event not found")
    return evt


@router.post("/dept/events/{event_id}/template", status_code=201)
async def upload_dept_event_template(
    event_id: str,
    template_file: UploadFile = File(...),
    current_user: User = Depends(require_role(UserRole.DEPT_COORDINATOR)),
):
    content_type = (template_file.content_type or "").lower()
    filename = (template_file.filename or "").lower()
    if not (
        content_type in ("image/png", "image/jpeg", "image/jpg")
        or filename.endswith((".png", ".jpg", ".jpeg"))
    ):
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Only PNG or JPEG files are accepted")

    department = _normalize_department(current_user.department)
    evt = await _get_dept_event_or_404(event_id, department, str(current_user.id))

    file_bytes = await template_file.read()
    try:
        img = Image.open(BytesIO(file_bytes))
        img.verify()
    except Exception:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Uploaded file is not a valid image")

    settings = get_settings()
    dept_slug = _slugify(department)
    event_slug = _slugify(evt.name)
    out_dir = settings.storage_root / "dept_templates" / dept_slug / str(evt.id)
    out_dir.mkdir(parents=True, exist_ok=True)

    out_name = f"template-{event_slug}-{datetime.utcnow().strftime('%Y%m%d%H%M%S')}-{uuid4().hex[:6]}.png"
    out_path = out_dir / out_name

    img = Image.open(BytesIO(file_bytes)).convert("RGBA")
    img.save(str(out_path), format="PNG")

    rel = out_path.resolve().relative_to(settings.storage_root.resolve()).as_posix()
    url = f"/storage/{rel}"

    # Mark previous templates for this event inactive.
    prev_templates = await DeptTemplate.find(
        DeptTemplate.department == department,
        DeptTemplate.event_id == evt.id,
        DeptTemplate.is_active == True,
    ).to_list()
    for t in prev_templates:
        await t.set({"is_active": False})

    doc = DeptTemplate(
        department=department,
        event_id=evt.id,
        original_filename=template_file.filename or out_name,
        template_path=str(out_path),
        template_url=url,
        is_active=True,
        uploaded_by_user_id=str(current_user.id),
    )
    await doc.insert()

    await evt.set({
        "template_id": doc.id,
        "preview_certificate_id": None,
        "preview_approved": False,
        "preview_approved_at": None,
        "preview_approved_by_user_id": None,
    })
    evt = await DeptEvent.get(evt.id)

    return {
        "message": "Template uploaded",
        "template": {
            "id": str(doc.id),
            "template_url": doc.template_url,
            "original_filename": doc.original_filename,
        },
        "event": _event_response(evt),
    }


@router.get("/dept/events/{event_id}/template")
async def get_dept_event_template(
    event_id: str,
    current_user: User = Depends(require_role(UserRole.DEPT_COORDINATOR)),
):
    department = _normalize_department(current_user.department)
    evt = await _get_dept_event_or_404(event_id, department, str(current_user.id))

    template_doc = None
    if evt.template_id:
        template_doc = await DeptTemplate.get(evt.template_id)

    if not template_doc:
        template_doc = await DeptTemplate.find_one(
            DeptTemplate.department == department,
            DeptTemplate.event_id == evt.id,
            DeptTemplate.is_active == True,
        )

    if not template_doc:
        return {"template": None}

    return {
        "template": {
            "id": str(template_doc.id),
            "template_url": template_doc.template_url,
            "original_filename": template_doc.original_filename,
        }
    }


@router.post("/dept/events/{event_id}/excel/headers")
async def extract_dept_event_excel_headers(
    event_id: str,
    excel_file: UploadFile = File(...),
    current_user: User = Depends(require_role(UserRole.DEPT_COORDINATOR)),
):
    department = _normalize_department(current_user.department)
    evt = await _get_dept_event_or_404(event_id, department, str(current_user.id))

    content = await excel_file.read()
    headers = _parse_excel_headers(content)
    return {"headers": headers}


@router.post("/dept/events/{event_id}/excel/preview")
async def preview_dept_event_excel_rows(
    event_id: str,
    excel_file: UploadFile = File(...),
    current_user: User = Depends(require_role(UserRole.DEPT_COORDINATOR)),
):
    department = _normalize_department(current_user.department)
    evt = await _get_dept_event_or_404(event_id, department, str(current_user.id))

    content = await excel_file.read()
    rows = _parse_excel_rows_dynamic(content)
    headers = _parse_excel_headers(content)

    preview = []
    for idx, row in enumerate(rows[:300], start=1):
        preview.append(
            {
                "id": str(idx),
                "name": (row.get("Name") or row.get("Student Name") or row.get("Participant") or "").strip(),
                "email": _pick_email_from_row(row) or "",
                "reg_no": (row.get("Reg No") or row.get("Registration Number") or row.get("Roll No") or "").strip(),
                "source": "excel",
                "verified": True,
                "raw": row,
            }
        )

    await evt.set({
        "excel_headers": headers,
        "excel_rows": rows,
        "excel_preview_row": rows[0] if rows else {},
        "excel_file_name": excel_file.filename or "",
        "excel_uploaded_at": datetime.utcnow(),
        "preview_certificate_id": None,
        "preview_approved": False,
        "preview_approved_at": None,
        "preview_approved_by_user_id": None,
    })

    return {
        "headers": headers,
        "preview_row": rows[0] if rows else {},
        "total_rows": len(rows),
        "preview_count": len(preview),
        "participants": preview,
    }


@router.get("/dept/events/{event_id}/mapping")
async def get_dept_event_mapping(
    event_id: str,
    current_user: User = Depends(require_role(UserRole.DEPT_COORDINATOR)),
):
    department = _normalize_department(current_user.department)
    evt = await _get_dept_event_or_404(event_id, department, str(current_user.id))

    return {
        "selected_fields": evt.selected_fields,
        "field_positions": evt.field_positions,
        "mapping_configured": evt.mapping_configured,
    }


@router.post("/dept/events/{event_id}/mapping")
async def save_dept_event_mapping(
    event_id: str,
    body: DeptEventFieldMappingRequest,
    current_user: User = Depends(require_role(UserRole.DEPT_COORDINATOR)),
):
    department = _normalize_department(current_user.department)
    evt = await _get_dept_event_or_404(event_id, department, str(current_user.id))

    selected = [str(f).strip() for f in body.selected_fields if str(f).strip()]
    if not selected:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Select at least one field")

    await evt.set({
        "selected_fields": selected,
        "field_positions": body.field_positions,
        "mapping_configured": True,
        "preview_certificate_id": None,
        "preview_approved": False,
        "preview_approved_at": None,
        "preview_approved_by_user_id": None,
    })
    evt = await DeptEvent.get(evt.id)
    return {
        "message": "Mapping saved",
        "selected_fields": evt.selected_fields,
        "field_positions": evt.field_positions,
        "mapping_configured": evt.mapping_configured,
    }


@router.get("/dept/events/{event_id}/certificates/preview")
async def get_dept_event_certificate_preview(
    event_id: str,
    current_user: User = Depends(require_role(UserRole.DEPT_COORDINATOR)),
):
    department = _normalize_department(current_user.department)
    evt = await _get_dept_event_or_404(event_id, department, str(current_user.id))

    preview_doc = None
    if evt.preview_certificate_id:
        preview_doc = await DeptCertificatePreview.get(evt.preview_certificate_id)

    if not preview_doc:
        return {
            "preview": None,
            "preview_approved": bool(evt.preview_approved),
            "source_rows_count": len(evt.excel_rows or []),
        }

    return {
        "preview": {
            "id": str(preview_doc.id),
            "cert_number": preview_doc.cert_number,
            "png_url": preview_doc.png_url,
            "participant_email": preview_doc.participant_email,
            "created_at": preview_doc.created_at,
            "approved": bool(preview_doc.approved),
            "approved_at": preview_doc.approved_at,
        },
        "preview_approved": bool(evt.preview_approved),
        "source_rows_count": len(evt.excel_rows or []),
    }


@router.post("/dept/events/{event_id}/certificates/preview")
async def generate_dept_event_certificate_preview(
    event_id: str,
    current_user: User = Depends(require_role(UserRole.DEPT_COORDINATOR)),
):
    department = _normalize_department(current_user.department)
    evt = await _get_dept_event_or_404(event_id, department, str(current_user.id))

    if not evt.template_id:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Upload event template first")
    if not evt.mapping_configured or not evt.selected_fields:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Configure field mapping first")
    if not evt.excel_rows:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Upload Excel and extract fields first")

    template_doc = await DeptTemplate.get(evt.template_id)
    if not template_doc or not Path(template_doc.template_path).exists():
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Template file not found for this event")

    asset = await _get_or_create_dept_asset(department)
    preview_row = evt.excel_preview_row or evt.excel_rows[0]
    dept_slug = _slugify(department)
    year = datetime.utcnow().year
    cert_number = f"PREVIEW-DPT-{dept_slug[:4].upper()}-{datetime.utcnow().strftime('%Y%m%d%H%M%S')}"

    png_bytes = _render_dept_certificate_dynamic(
        template_path=Path(template_doc.template_path),
        row=preview_row,
        cert_number=cert_number,
        selected_fields=evt.selected_fields,
        field_positions=evt.field_positions,
        logo_path=asset.logo_path,
        sig1_path=asset.signature1_path,
        event_date=evt.event_date,
    )
    png_url = save_cert_png(png_bytes, dept_slug, year, cert_number)

    preview_doc = DeptCertificatePreview(
        department=department,
        coordinator_user_id=str(current_user.id),
        event_id=evt.id,
        template_id=template_doc.id,
        cert_number=cert_number,
        participant_email=_pick_email_from_row(preview_row),
        preview_row=preview_row,
        png_url=png_url,
        approved=False,
    )
    await preview_doc.insert()

    await evt.set({
        "preview_certificate_id": preview_doc.id,
        "preview_approved": False,
        "preview_approved_at": None,
        "preview_approved_by_user_id": None,
    })

    return {
        "message": "Preview certificate generated",
        "preview": {
            "id": str(preview_doc.id),
            "cert_number": preview_doc.cert_number,
            "png_url": preview_doc.png_url,
            "participant_email": preview_doc.participant_email,
            "created_at": preview_doc.created_at,
            "approved": False,
        },
    }


@router.post("/dept/events/{event_id}/certificates/preview/approve")
async def approve_dept_event_certificate_preview(
    event_id: str,
    current_user: User = Depends(require_role(UserRole.DEPT_COORDINATOR)),
):
    department = _normalize_department(current_user.department)
    evt = await _get_dept_event_or_404(event_id, department, str(current_user.id))

    if not evt.preview_certificate_id:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Generate preview certificate first")

    preview_doc = await DeptCertificatePreview.get(evt.preview_certificate_id)
    if not preview_doc:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Preview certificate not found")

    now = datetime.utcnow()
    preview_doc.approved = True
    preview_doc.approved_at = now
    preview_doc.approved_by_user_id = str(current_user.id)
    await preview_doc.save()

    await evt.set({
        "preview_approved": True,
        "preview_approved_at": now,
        "preview_approved_by_user_id": str(current_user.id),
    })

    return {
        "message": "Preview approved. You can now generate certificates for this event.",
        "preview_id": str(preview_doc.id),
        "approved_at": now,
    }


@router.post("/dept/events/{event_id}/certificates/generate")
async def generate_dept_event_certificates(
    event_id: str,
    current_user: User = Depends(require_role(UserRole.DEPT_COORDINATOR)),
):
    department = _normalize_department(current_user.department)
    evt = await _get_dept_event_or_404(event_id, department, str(current_user.id))

    if not evt.template_id:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Upload event template first")
    if not evt.mapping_configured or not evt.selected_fields:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Configure field mapping first")
    if not evt.excel_rows:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Upload Excel and extract fields first")
    if not evt.preview_approved:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Approve preview certificate before generating")

    template_doc = await DeptTemplate.get(evt.template_id)
    if not template_doc or not Path(template_doc.template_path).exists():
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Template file not found for this event")

    asset = await _get_or_create_dept_asset(department)
    rows = evt.excel_rows

    dept_slug = _slugify(department)
    year = datetime.utcnow().year
    timestamp = datetime.utcnow().strftime("%Y%m%d%H%M%S")

    generated = 0
    skipped = 0
    cert_numbers: list[str] = []
    skipped_emails: list[str] = []
    participant_keys: set[str] = set()

    existing_certs = await DeptCertificate.find(
        DeptCertificate.department == department,
        DeptCertificate.event_id == str(evt.id),
    ).to_list()
    existing_emails = {
        str(c.participant_email).strip().lower()
        for c in existing_certs
        if c.participant_email and str(c.participant_email).strip()
    }
    seen_upload_emails: set[str] = set()

    def _normalize_email(value: Optional[str]) -> str:
        return (value or "").strip().lower()

    for idx, row in enumerate(rows, start=1):
        participant_email = _normalize_email(_pick_email_from_row(row))
        if not participant_email:
            skipped += 1
            skipped_emails.append("<missing-email>")
            continue

        if participant_email in existing_emails or participant_email in seen_upload_emails:
            skipped += 1
            skipped_emails.append(participant_email)
            continue

        existing_for_email = await DeptCertificate.find_one({
            "department": department,
            "event_id": str(evt.id),
            "participant_email": {
                "$regex": f"^{re.escape(participant_email)}$",
                "$options": "i",
            },
        })
        if existing_for_email:
            skipped += 1
            skipped_emails.append(participant_email)
            existing_emails.add(participant_email)
            continue

        seen_upload_emails.add(participant_email)

        cert_number = f"DPT-{dept_slug[:4].upper()}-{timestamp}-{idx:03d}"
        png_bytes = _render_dept_certificate_dynamic(
            template_path=Path(template_doc.template_path),
            row=row,
            cert_number=cert_number,
            selected_fields=evt.selected_fields,
            field_positions=evt.field_positions,
            logo_path=asset.logo_path,
            sig1_path=asset.signature1_path,
            event_date=evt.event_date,
        )
        png_url = save_cert_png(png_bytes, dept_slug, year, cert_number)

        name = (row.get("Name") or row.get("Student Name") or "").strip() or "Unknown"
        class_name = (row.get("Class") or row.get("Department") or row.get("Semester") or "").strip() or "-"
        contribution = (row.get("Contribution") or row.get("Role") or row.get("Participation") or "").strip() or "Participant"

        doc = DeptCertificate(
            cert_number=cert_number,
            department=department,
            coordinator_user_id=str(current_user.id),
            event_id=str(evt.id),
            name=name,
            class_name=class_name,
            contribution=contribution,
            participant_email=participant_email,
            png_url=png_url,
            created_at=datetime.utcnow(),
        )
        await doc.insert()

        existing_emails.add(participant_email)

        participant_keys.add(f"{name.lower()}|{class_name.lower()}")
        generated += 1
        cert_numbers.append(cert_number)

    await evt.set({
        "participant_count": len(participant_keys),
        "cert_count": generated,
        "status": DeptEventStatus.ACTIVE,
    })

    return {
        "generated": generated,
        "skipped": skipped,
        "total_rows": len(rows),
        "message": f"Generated {generated} certificate(s) for event; skipped {skipped} duplicate email row(s).",
        "cert_numbers": cert_numbers,
        "skipped_emails": skipped_emails,
    }


@router.get("/dept/events/{event_id}/certificates")
async def list_dept_event_certificates(
    event_id: str,
    current_user: User = Depends(require_role(UserRole.DEPT_COORDINATOR)),
):
    department = _normalize_department(current_user.department)
    evt = await _get_dept_event_or_404(event_id, department, str(current_user.id))

    certs = await DeptCertificate.find(
        DeptCertificate.department == department,
        DeptCertificate.event_id == str(evt.id),
    ).sort(-DeptCertificate.created_at).to_list()

    return [
        {
            "id": str(c.id),
            "cert_number": c.cert_number,
            "participant_name": c.name,
            "participant_email": c.participant_email,
            "cert_type": "participant",
            "status": "emailed" if c.emailed_at else ("failed" if c.email_error else "generated"),
            "generated_at": c.created_at,
            "emailed_at": c.emailed_at,
            "failure_reason": c.email_error,
            "png_url": c.png_url,
        }
        for c in certs
    ]


@router.post("/dept/events/{event_id}/certificates/send")
async def send_dept_event_certificates(
    event_id: str,
    current_user: User = Depends(require_role(UserRole.DEPT_COORDINATOR)),
):
    department = _normalize_department(current_user.department)
    evt = await _get_dept_event_or_404(event_id, department, str(current_user.id))

    certs = await DeptCertificate.find(
        DeptCertificate.department == department,
        DeptCertificate.event_id == str(evt.id),
    ).to_list()

    sendable = [c for c in certs if not c.emailed_at and c.participant_email and c.png_url]
    if not sendable:
        return {"queued": 0, "sent": 0, "failed": 0, "message": "No pending certificates to send."}

    sent = 0
    failed = 0
    for cert in sendable:
        local_path = storage_url_to_path(cert.png_url)
        if not local_path or not Path(local_path).exists():
            cert.email_error = "Certificate file not found"
            await cert.save()
            failed += 1
            continue

        ok = await send_certificate_email(
            recipient_email=cert.participant_email,
            recipient_name=cert.name,
            cert_number=cert.cert_number,
            event_name=evt.name,
            club_name=department,
            png_path=local_path,
        )
        if ok:
            cert.emailed_at = datetime.utcnow()
            cert.email_error = None
            sent += 1
        else:
            cert.email_error = "Email delivery failed"
            failed += 1
        await cert.save()

    return {
        "queued": len(sendable),
        "sent": sent,
        "failed": failed,
        "message": f"Sent {sent} email(s); {failed} failed.",
    }


@router.post("/dept/events/{event_id}/certificates/{cert_id}/send")
async def send_single_dept_event_certificate(
    event_id: str,
    cert_id: str,
    current_user: User = Depends(require_role(UserRole.DEPT_COORDINATOR)),
):
    department = _normalize_department(current_user.department)
    evt = await _get_dept_event_or_404(event_id, department, str(current_user.id))

    cert = await DeptCertificate.get(cert_id)
    if not cert or cert.department != department or cert.event_id != str(evt.id):
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Certificate not found for this event")

    if not cert.participant_email:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "No recipient email available for this certificate")

    local_path = storage_url_to_path(cert.png_url)
    if not local_path or not Path(local_path).exists():
        cert.email_error = "Certificate file not found"
        await cert.save()
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Certificate file not found")

    ok = await send_certificate_email(
        recipient_email=cert.participant_email,
        recipient_name=cert.name,
        cert_number=cert.cert_number,
        event_name=evt.name,
        club_name=department,
        png_path=local_path,
    )

    if ok:
        cert.emailed_at = datetime.utcnow()
        cert.email_error = None
        await cert.save()
        return {"message": "Certificate email sent successfully."}

    cert.email_error = "Email delivery failed"
    await cert.save()
    raise HTTPException(status.HTTP_500_INTERNAL_SERVER_ERROR, "Failed to send certificate email")


@router.get("/dept/certificates/assets-status")
async def get_asset_status(
    current_user: User = Depends(require_role(UserRole.DEPT_COORDINATOR)),
):
    department = _normalize_department(current_user.department)
    asset = await _get_or_create_dept_asset(department)

    template_url = asset.certificate_template_url
    if not template_url:
        try:
            fallback_path = _pick_single_template()
            template_url = f"/static/certificate_templates/{fallback_path.name}"
        except Exception:
            template_url = None

    return {
        "department": department,
        "has_logo": bool(asset.logo_path),
        "has_signature_primary": bool(asset.signature1_path),
        "has_signature_secondary": bool(asset.signature2_path),
        "has_template": bool(asset.certificate_template_path),
        "template_url": template_url,
        "positions_configured": asset.positions_configured,
    }


@router.post("/dept/certificates/template/upload")
async def upload_certificate_template(
    template_file: UploadFile = File(...),
    current_user: User = Depends(require_role(UserRole.DEPT_COORDINATOR)),
):
    """Upload a custom PNG/JPEG certificate template for this department.

    Saves to storage/dept_templates/{dept_slug}/template.png and persists
    the path on DeptAsset so subsequent generation uses this template instead
    of the global system template.
    """
    # Validate file type
    content_type = (template_file.content_type or "").lower()
    filename = (template_file.filename or "").lower()
    if not (
        content_type in ("image/png", "image/jpeg", "image/jpg")
        or filename.endswith((".png", ".jpg", ".jpeg"))
    ):
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            "Only PNG or JPEG files are accepted for certificate templates",
        )

    department = _normalize_department(current_user.department)
    dept_slug = _slugify(department)

    # Determine storage path
    from ...config import get_settings
    settings = get_settings()
    template_dir = Path(settings.storage_root) / "dept_templates" / dept_slug
    template_dir.mkdir(parents=True, exist_ok=True)
    template_path = template_dir / "template.png"

    # Read, validate it opens as an image, and save
    file_bytes = await template_file.read()
    try:
        img = Image.open(BytesIO(file_bytes))
        img.verify()  # fast check — raises if not a valid image
    except Exception:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            "Uploaded file could not be opened as an image. Ensure it is a valid PNG or JPEG.",
        )

    # Re-open after verify (verify() exhausts the file pointer)
    img = Image.open(BytesIO(file_bytes)).convert("RGBA")
    img.save(str(template_path), format="PNG")

    # Persist path on DeptAsset
    asset = await _get_or_create_dept_asset(department)
    asset.certificate_template_path = str(template_path)
    asset.certificate_template_url = f"/storage/dept_templates/{dept_slug}/template.png"
    asset.updated_at = datetime.utcnow()
    await asset.save()

    return {
        "message": "Certificate template uploaded successfully",
        "template_url": asset.certificate_template_url,
        "department": department,
    }


@router.post("/dept/certificates/generate")
async def generate_department_certificates(
    excel_file: UploadFile = File(...),
    logo_file: UploadFile | None = File(None),
    signature_primary_file: UploadFile | None = File(None),
    signature_secondary_file: UploadFile | None = File(None),
    current_user: User = Depends(require_role(UserRole.DEPT_COORDINATOR)),
):
    department = _normalize_department(current_user.department)
    dept_slug = _slugify(department)
    year = datetime.utcnow().year

    asset = await _get_or_create_dept_asset(department)

    if logo_file is not None:
        logo_bytes = await logo_file.read()
        asset.logo_path = save_logo(logo_bytes, dept_slug)
        asset.logo_hash = hashlib.md5(logo_bytes).hexdigest()
        asset.logo_url = storage_path_to_url(asset.logo_path)

    if signature_primary_file is not None:
        sig1_bytes = await signature_primary_file.read()
        asset.signature1_path = process_signature(sig1_bytes, dept_slug)
        asset.signature1_hash = hashlib.md5(sig1_bytes).hexdigest()
        asset.signature1_url = storage_path_to_url(asset.signature1_path)

    if signature_secondary_file is not None:
        sig2_bytes = await signature_secondary_file.read()
        asset.signature2_path = process_signature(sig2_bytes, dept_slug)
        asset.signature2_hash = hashlib.md5(sig2_bytes).hexdigest()
        asset.signature2_url = storage_path_to_url(asset.signature2_path)

    if not asset.logo_path or not asset.signature1_path or not asset.signature2_path:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            "Logo, primary signature, and secondary signature are required the first time",
        )

    asset.updated_at = datetime.utcnow()
    await asset.save()

    excel_bytes = await excel_file.read()
    rows = _parse_dept_excel(excel_bytes)

    # Prefer the department's custom template; fall back to the global system template.
    if asset.certificate_template_path and Path(asset.certificate_template_path).exists():
        template_path = Path(asset.certificate_template_path)
    else:
        template_path = _pick_single_template()


    generated = 0
    cert_numbers: list[str] = []
    timestamp = datetime.utcnow().strftime("%Y%m%d%H%M%S")

    # Get current field positions for rendering
    field_positions = None
    if asset.field_positions:
        field_positions = { k: v.model_dump() for k, v in asset.field_positions.items() }

    for idx, row in enumerate(rows, start=1):
        cert_number = f"DPT-{dept_slug[:4].upper()}-{timestamp}-{idx:03d}"
        png_bytes = _render_dept_certificate(
            template_path=template_path,
            name=row["name"],
            class_name=row["class_name"],
            contribution=row["contribution"],
            cert_number=cert_number,
            logo_path=asset.logo_path,
            sig1_path=asset.signature1_path,
            sig2_path=asset.signature2_path,
            field_positions=field_positions,
        )
        png_url = save_cert_png(png_bytes, dept_slug, year, cert_number)

        doc = DeptCertificate(
            cert_number=cert_number,
            department=department,
            coordinator_user_id=str(current_user.id),
            name=row["name"],
            class_name=row["class_name"],
            contribution=row["contribution"],
            png_url=png_url,
            created_at=datetime.utcnow(),
        )
        await doc.insert()
        generated += 1
        cert_numbers.append(cert_number)

    return {
        "generated": generated,
        "total_rows": len(rows),
        "message": f"Generated {generated} department certificate(s)",
        "cert_numbers": cert_numbers,
    }


@router.post("/dept/certificates/download-zip")
async def download_department_certificates_zip(
    body: DeptBatchDownloadRequest,
    current_user: User = Depends(require_role(UserRole.DEPT_COORDINATOR)),
):
    """Zip selected certificate PNGs for the current generation batch."""
    if not body.cert_numbers:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "cert_numbers is required")

    from zipfile import ZipFile, ZIP_DEFLATED
    import io
    from ...services.storage_service import storage_url_to_path

    department = _normalize_department(current_user.department)
    certs = await DeptCertificate.find(
        {
            "department": department,
            "cert_number": {"$in": body.cert_numbers},
        }
    ).to_list()

    if not certs:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "No certificates found for the requested batch")

    cert_map = {c.cert_number: c for c in certs}
    buf = io.BytesIO()
    added_count = 0
    with ZipFile(buf, "w", ZIP_DEFLATED) as zf:
        for cert_number in body.cert_numbers:
            cert = cert_map.get(cert_number)
            if not cert or not cert.png_url:
                continue
            local_path = storage_url_to_path(cert.png_url)
            if local_path and Path(local_path).exists():
                zf.write(local_path, f"{cert.cert_number}.png")
                added_count += 1

    if added_count == 0:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "No certificate PNG files found for this batch")

    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/zip",
        headers={"Content-Disposition": "attachment; filename=dept_certs_batch.zip"},
    )


@router.get("/dept/certificates")
async def list_department_generated_certificates(
    limit: int = Query(50, ge=1, le=200),
    current_user: User = Depends(require_role(UserRole.DEPT_COORDINATOR)),
):
    department = _normalize_department(current_user.department)

    certs = await DeptCertificate.find(
        DeptCertificate.department == department
    ).sort(-DeptCertificate.created_at).limit(limit).to_list()

    return [
        {
            "id": str(c.id),
            "cert_number": c.cert_number,
            "name": c.name,
            "class_name": c.class_name,
            "contribution": c.contribution,
            "issued_at": c.created_at,
            "png_url": c.png_url,
        }
        for c in certs
    ]


@router.get("/dept/certificates/download-all")
async def download_all_department_certificates(
    current_user: User = Depends(require_role(UserRole.DEPT_COORDINATOR)),
):
    """Zips all generated certificate PNGs for the user's department."""
    from zipfile import ZipFile, ZIP_DEFLATED
    import io
    from ...services.storage_service import storage_url_to_path
    
    department = _normalize_department(current_user.department)
    
    certs = await DeptCertificate.find(
        DeptCertificate.department == department
    ).to_list()
    
    if not certs:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "No certificates found in your department to download")
        
    buf = io.BytesIO()
    added_count = 0
    with ZipFile(buf, "w", ZIP_DEFLATED) as zf:
        for c in certs:
            if not c.png_url:
                continue
            local_path = storage_url_to_path(c.png_url)
            if local_path and Path(local_path).exists():
                zf.write(local_path, f"{c.cert_number}.png")
                added_count += 1
                
    if added_count == 0:
        raise HTTPException(
            status.HTTP_404_NOT_FOUND, 
            "No certificate PNG files found on disk for this department"
        )
        
    buf.seek(0)
    dept_slug = _slugify(department)
    return StreamingResponse(
        buf,
        media_type="application/zip",
        headers={"Content-Disposition": f'attachment; filename="{dept_slug}_certificates.zip"'}
    )


@router.get("/dept/certificates/field-positions")
async def get_field_positions(
    current_user: User = Depends(require_role(UserRole.DEPT_COORDINATOR)),
):
    """Retrieve saved field position configurations for the department"""
    department = _normalize_department(current_user.department)
    asset = await _get_or_create_dept_asset(department)
    
    return {
        "field_positions": asset.field_positions,
        "positions_configured": asset.positions_configured,
    }


@router.post("/dept/certificates/field-positions")
async def configure_field_positions(
    positions: dict,
    current_user: User = Depends(require_role(UserRole.DEPT_COORDINATOR)),
):
    """Save field position configurations for department certificates"""
    from ...models.dept_asset import DeptFieldPosition
    
    department = _normalize_department(current_user.department)
    asset = await _get_or_create_dept_asset(department)
    
    # Validate and convert positions to DeptFieldPosition objects
    validated_positions = {}
    for field_name, pos_data in positions.items():
        try:
            # Create DeptFieldPosition from the position data
            field_pos = DeptFieldPosition(
                field_name=field_name,
                x_percent=float(pos_data.get("x_percent", 0)),
                y_percent=float(pos_data.get("y_percent", 0)),
                font_size=int(pos_data.get("font_size", 24)),
            )
            validated_positions[field_name] = field_pos
        except (ValueError, TypeError) as e:
            raise HTTPException(
                status.HTTP_400_BAD_REQUEST,
                f"Invalid position data for field '{field_name}': {str(e)}",
            )
    
    asset.field_positions = validated_positions
    asset.positions_configured = True
    asset.updated_at = datetime.utcnow()
    await asset.save()
    
    return {
        "message": "Field positions configured successfully",
        "field_positions": asset.field_positions,
        "positions_configured": asset.positions_configured,
    }


@router.get("/dept/students")
async def list_department_students(
    batch: Optional[str] = None,
    sort_by: str = Query("total_credits", pattern="^(total_credits|name)$"),
    order: str = Query("desc", pattern="^(asc|desc)$"),
    current_user: User = Depends(require_role(UserRole.DEPT_COORDINATOR)),
):
    query = {"department": current_user.department}
    if batch:
        query["batch"] = batch

    students = await StudentCredit.find(query).to_list()

    reverse = order == "desc"
    if sort_by == "total_credits":
        students.sort(key=lambda s: s.total_credits, reverse=reverse)
    else:
        students.sort(key=lambda s: (s.student_name or "").lower(), reverse=reverse)

    return [
        {
            "id": str(s.id),
            "student_email": s.student_email,
            "registration_number": s.registration_number,
            "student_name": s.student_name,
            "department": s.department,
            "batch": s.batch,
            "section": s.section,
            "total_credits": s.total_credits,
            "last_updated": s.last_updated,
        }
        for s in students
    ]


@router.get("/dept/students/{student_email}")
async def get_student_detail(
    student_email: str,
    current_user: User = Depends(require_role(UserRole.DEPT_COORDINATOR)),
):
    student = await StudentCredit.find_one(
        StudentCredit.student_email == student_email,
        StudentCredit.department == current_user.department,
    )
    if not student:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Student not found in your department")

    return {
        "student_name": student.student_name,
        "student_email": student.student_email,
        "registration_number": student.registration_number,
        "department": student.department,
        "batch": student.batch,
        "section": student.section,
        "total_credits": student.total_credits,
        "credit_history": [e.model_dump() for e in student.credit_history],
    }


@router.get("/dept/export")
async def export_department(
    current_user: User = Depends(require_role(UserRole.DEPT_COORDINATOR)),
):
    students = await StudentCredit.find(
        StudentCredit.department == current_user.department
    ).to_list()

    wb = Workbook()
    ws = wb.active
    ws.title = "Student Credits"
    headers = ["Name", "Email", "Reg No", "Section", "Batch", "Total Credits", "Last Activity"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    for row_idx, s in enumerate(students, 2):
        last_activity = s.credit_history[-1].awarded_at if s.credit_history else None
        ws.cell(row=row_idx, column=1, value=s.student_name)
        ws.cell(row=row_idx, column=2, value=s.student_email)
        ws.cell(row=row_idx, column=3, value=s.registration_number)
        ws.cell(row=row_idx, column=4, value=s.section)
        ws.cell(row=row_idx, column=5, value=s.batch)
        ws.cell(row=row_idx, column=6, value=s.total_credits)
        ws.cell(row=row_idx, column=7, value=str(last_activity) if last_activity else "")

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    dept = _slugify(current_user.department or "dept")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={dept}_credits.xlsx"},
    )
