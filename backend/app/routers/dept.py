from io import BytesIO
from pathlib import Path
from typing import Optional
import hashlib
import re
from datetime import datetime

from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, UploadFile, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from PIL import Image, ImageDraw, ImageFont

from ..core.dependencies import require_role
from ..models.user import User, UserRole
from ..models.student_credit import StudentCredit
from ..models.dept_asset import DeptAsset
from ..models.dept_certificate import DeptCertificate
from ..services.signature_service import process_signature, save_logo
from ..services.storage_service import save_cert_png

router = APIRouter(tags=["Department"])


def _slugify(value: str) -> str:
    text = re.sub(r"[^a-zA-Z0-9]+", "-", value.strip().lower())
    return text.strip("-") or "dept"


def _normalize_department(department: Optional[str]) -> str:
    if not department:
        return "General"
    return department.strip()


def _pick_single_template() -> Path:
    templates_dir = Path(__file__).parent.parent / "static" / "certificate_templates"
    pngs = sorted(templates_dir.glob("*.png"))
    if not pngs:
        raise HTTPException(status.HTTP_500_INTERNAL_SERVER_ERROR, "No PNG certificate templates available")
    return pngs[0]


def _parse_dept_excel(file_bytes: bytes) -> list[dict]:
    wb = load_workbook(filename=BytesIO(file_bytes), read_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    headers_row = next(rows, None)
    if not headers_row:
        wb.close()
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Excel is empty")

    headers = [str(h).strip() if h is not None else "" for h in headers_row]
    norm = [h.lower().replace(" ", "").replace("_", "") for h in headers]

    def idx_for(*names: str) -> int:
        for i, key in enumerate(norm):
            if key in names:
                return i
        return -1

    name_i = idx_for("name")
    class_i = idx_for("class", "classname", "departmentclass")
    contribution_i = idx_for("contribution", "whatcontribution", "role")

    if name_i < 0 or class_i < 0 or contribution_i < 0:
        wb.close()
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            "Excel must include columns: Name, Class, Contribution",
        )

    parsed = []
    for row in rows:
        if not row or all(v is None for v in row):
            continue

        name = str(row[name_i]).strip() if name_i < len(row) and row[name_i] is not None else ""
        class_name = str(row[class_i]).strip() if class_i < len(row) and row[class_i] is not None else ""
        contribution = str(row[contribution_i]).strip() if contribution_i < len(row) and row[contribution_i] is not None else ""

        if not name or not class_name or not contribution:
            continue

        parsed.append(
            {
                "name": name,
                "class_name": class_name,
                "contribution": contribution,
            }
        )

    wb.close()
    if not parsed:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "No valid rows found in Excel")
    return parsed


def _load_font(size: int):
    font_path = Path(__file__).parent.parent / "static" / "fonts" / "Montserrat-Bold.ttf"
    try:
        if font_path.exists():
            return ImageFont.truetype(str(font_path), size)
    except Exception:
        pass
    return ImageFont.load_default()


def _render_dept_certificate(
    template_path: Path,
    name: str,
    class_name: str,
    contribution: str,
    cert_number: str,
    logo_path: Optional[str],
    sig1_path: Optional[str],
    sig2_path: Optional[str],
) -> bytes:
    img = Image.open(str(template_path)).convert("RGBA")
    draw = ImageDraw.Draw(img)
    w, h = img.size

    name_font = _load_font(max(28, int(w * 0.028)))
    info_font = _load_font(max(22, int(w * 0.018)))
    cert_font = _load_font(max(16, int(w * 0.013)))

    draw.text((w * 0.5, h * 0.46), name, fill=(28, 35, 70, 255), font=name_font, anchor="mm")
    draw.text((w * 0.5, h * 0.56), f"Class: {class_name}", fill=(45, 45, 45, 255), font=info_font, anchor="mm")
    draw.text((w * 0.5, h * 0.64), f"Contribution: {contribution}", fill=(45, 45, 45, 255), font=info_font, anchor="mm")
    draw.text((w * 0.82, h * 0.05), cert_number, fill=(44, 61, 127, 255), font=cert_font, anchor="lm")

    def _paste_scaled(path: Optional[str], cx: float, cy: float, max_w_ratio: float, max_h_ratio: float):
        if not path:
            return
        p = Path(path)
        if not p.exists():
            return
        try:
            asset = Image.open(str(p)).convert("RGBA")
            max_w = int(w * max_w_ratio)
            max_h = int(h * max_h_ratio)
            asset.thumbnail((max_w, max_h), Image.LANCZOS)
            x = int(w * cx - asset.width / 2)
            y = int(h * cy - asset.height / 2)
            img.paste(asset, (x, y), asset.split()[3])
        except Exception:
            return

    _paste_scaled(logo_path, 0.12, 0.12, 0.16, 0.16)
    _paste_scaled(sig1_path, 0.18, 0.84, 0.2, 0.1)
    _paste_scaled(sig2_path, 0.78, 0.84, 0.2, 0.1)

    out = BytesIO()
    if img.mode == "RGBA":
        bg = Image.new("RGB", img.size, "white")
        bg.paste(img, mask=img.split()[3])
        bg.save(out, format="PNG")
    else:
        img.save(out, format="PNG")
    out.seek(0)
    return out.getvalue()


async def _get_or_create_dept_asset(department: str) -> DeptAsset:
    asset = await DeptAsset.find_one(DeptAsset.department == department)
    if asset:
        return asset

    asset = DeptAsset(department=department)
    await asset.insert()
    return asset


@router.get("/dept/certificates/assets-status")
async def get_asset_status(
    current_user: User = Depends(require_role(UserRole.DEPT_COORDINATOR)),
):
    department = _normalize_department(current_user.department)
    asset = await _get_or_create_dept_asset(department)

    return {
        "department": department,
        "has_logo": bool(asset.logo_path),
        "has_signature_primary": bool(asset.signature1_path),
        "has_signature_secondary": bool(asset.signature2_path),
    }


@router.post("/dept/certificates/generate")
async def generate_department_certificates(
    excel_file: UploadFile = File(...),
    logo_file: UploadFile | None = File(None),
    signature_primary_file: UploadFile | None = File(None),
    signature_secondary_file: UploadFile | None = File(None),
    current_user: User = Depends(require_role(UserRole.DEPT_COORDINATOR)),
):
    department = _normalize_department(current_user.department)
    dept_slug = _slugify(department)
    year = datetime.utcnow().year

    asset = await _get_or_create_dept_asset(department)

    if logo_file is not None:
        logo_bytes = await logo_file.read()
        asset.logo_path = save_logo(logo_bytes, dept_slug)
        asset.logo_hash = hashlib.md5(logo_bytes).hexdigest()
        asset.logo_url = asset.logo_path

    if signature_primary_file is not None:
        sig1_bytes = await signature_primary_file.read()
        asset.signature1_path = process_signature(sig1_bytes, dept_slug)
        asset.signature1_hash = hashlib.md5(sig1_bytes).hexdigest()
        asset.signature1_url = asset.signature1_path

    if signature_secondary_file is not None:
        sig2_bytes = await signature_secondary_file.read()
        asset.signature2_path = process_signature(sig2_bytes, dept_slug)
        asset.signature2_hash = hashlib.md5(sig2_bytes).hexdigest()
        asset.signature2_url = asset.signature2_path

    if not asset.logo_path or not asset.signature1_path or not asset.signature2_path:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            "Logo, primary signature, and secondary signature are required the first time",
        )

    asset.updated_at = datetime.utcnow()
    await asset.save()

    excel_bytes = await excel_file.read()
    rows = _parse_dept_excel(excel_bytes)
    template_path = _pick_single_template()

    generated = 0
    timestamp = datetime.utcnow().strftime("%Y%m%d%H%M%S")

    for idx, row in enumerate(rows, start=1):
        cert_number = f"DPT-{dept_slug[:4].upper()}-{timestamp}-{idx:03d}"
        png_bytes = _render_dept_certificate(
            template_path=template_path,
            name=row["name"],
            class_name=row["class_name"],
            contribution=row["contribution"],
            cert_number=cert_number,
            logo_path=asset.logo_path,
            sig1_path=asset.signature1_path,
            sig2_path=asset.signature2_path,
        )
        png_url = save_cert_png(png_bytes, dept_slug, year, cert_number)

        doc = DeptCertificate(
            cert_number=cert_number,
            department=department,
            coordinator_user_id=str(current_user.id),
            name=row["name"],
            class_name=row["class_name"],
            contribution=row["contribution"],
            png_url=png_url,
            created_at=datetime.utcnow(),
        )
        await doc.insert()
        generated += 1

    return {
        "generated": generated,
        "total_rows": len(rows),
        "message": f"Generated {generated} department certificate(s)",
    }


@router.get("/dept/certificates")
async def list_department_generated_certificates(
    limit: int = Query(50, ge=1, le=200),
    current_user: User = Depends(require_role(UserRole.DEPT_COORDINATOR)),
):
    department = _normalize_department(current_user.department)

    certs = await DeptCertificate.find(
        DeptCertificate.department == department
    ).sort(-DeptCertificate.created_at).limit(limit).to_list()

    return [
        {
            "id": str(c.id),
            "cert_number": c.cert_number,
            "name": c.name,
            "class_name": c.class_name,
            "contribution": c.contribution,
            "issued_at": c.created_at,
            "png_url": c.png_url,
        }
        for c in certs
    ]


@router.get("/dept/students")
async def list_department_students(
    batch: Optional[str] = None,
    sort_by: str = Query("total_credits", pattern="^(total_credits|name)$"),
    order: str = Query("desc", pattern="^(asc|desc)$"),
    current_user: User = Depends(require_role(UserRole.DEPT_COORDINATOR)),
):
    query = {"department": current_user.department}
    if batch:
        query["batch"] = batch

    students = await StudentCredit.find(query).to_list()

    reverse = order == "desc"
    if sort_by == "total_credits":
        students.sort(key=lambda s: s.total_credits, reverse=reverse)
    else:
        students.sort(key=lambda s: (s.student_name or "").lower(), reverse=reverse)

    return [
        {
            "id": str(s.id),
            "student_email": s.student_email,
            "registration_number": s.registration_number,
            "student_name": s.student_name,
            "department": s.department,
            "batch": s.batch,
            "section": s.section,
            "total_credits": s.total_credits,
            "last_updated": s.last_updated,
        }
        for s in students
    ]


@router.get("/dept/students/{student_email}")
async def get_student_detail(
    student_email: str,
    current_user: User = Depends(require_role(UserRole.DEPT_COORDINATOR)),
):
    student = await StudentCredit.find_one(
        StudentCredit.student_email == student_email,
        StudentCredit.department == current_user.department,
    )
    if not student:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Student not found in your department")

    return {
        "student_name": student.student_name,
        "student_email": student.student_email,
        "registration_number": student.registration_number,
        "department": student.department,
        "batch": student.batch,
        "section": student.section,
        "total_credits": student.total_credits,
        "credit_history": [e.model_dump() for e in student.credit_history],
    }


@router.get("/dept/export")
async def export_department(
    current_user: User = Depends(require_role(UserRole.DEPT_COORDINATOR)),
):
    students = await StudentCredit.find(
        StudentCredit.department == current_user.department
    ).to_list()

    wb = Workbook()
    ws = wb.active
    ws.title = "Student Credits"
    headers = ["Name", "Email", "Reg No", "Section", "Batch", "Total Credits", "Last Activity"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    for row_idx, s in enumerate(students, 2):
        last_activity = s.credit_history[-1].awarded_at if s.credit_history else None
        ws.cell(row=row_idx, column=1, value=s.student_name)
        ws.cell(row=row_idx, column=2, value=s.student_email)
        ws.cell(row=row_idx, column=3, value=s.registration_number)
        ws.cell(row=row_idx, column=4, value=s.section)
        ws.cell(row=row_idx, column=5, value=s.batch)
        ws.cell(row=row_idx, column=6, value=s.total_credits)
        ws.cell(row=row_idx, column=7, value=str(last_activity) if last_activity else "")

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    dept = _slugify(current_user.department or "dept")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={dept}_credits.xlsx"},
    )
