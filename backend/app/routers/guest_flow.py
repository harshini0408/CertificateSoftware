"""
Guest Flow Router — 5-step wizard backend.

All routes are prefixed: /clubs/{club_id}/events/{event_id}/guest
All routes are protected by require_event_access (guest role check already
verifies club_id + event_id match the JWT).

Step 1  POST /template         — Upload custom PNG background template
Step 2  POST /excel            — Upload XLSX, parse headers + data
        POST /config           — Save selected_columns + email_column
Step 3  POST /field-positions  — Reused from image_templates.py
Step 4  POST /generate         — Pillow batch generation
Step 5  GET  /zip              — Download all generated PNGs as ZIP
        POST /send-emails      — Email each certificate to the email column
        GET  /status           — Poll current guest event state
"""

import asyncio
import io
import logging
import uuid
import zipfile
from pathlib import Path
from typing import Dict, List, Optional

from beanie import PydanticObjectId
from fastapi import APIRouter, Depends, File, HTTPException, UploadFile, status
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

from ..config import get_settings
from ..core.dependencies import require_event_access
from ..models.event import Event
from ..models.field_position import FieldPosition
from ..models.user import User
from ..services.email_service import send_certificate_email
from ..services.storage_service import storage_path_to_url

logger = logging.getLogger(__name__)
settings = get_settings()

router = APIRouter(
    prefix="/clubs/{club_id}/events/{event_id}/guest",
    tags=["Guest Flow"],
)

# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

def _guest_certs_dir(club_id: str, event_id: str) -> Path:
    """Directory where generated guest certificates are stored."""
    d = settings.certs_dir / "guest" / str(club_id) / str(event_id)
    d.mkdir(parents=True, exist_ok=True)
    return d

def _guest_templates_dir() -> Path:
    """Directory where guest-uploaded PNG templates are stored."""
    d = settings.storage_root / "guest_templates"
    d.mkdir(parents=True, exist_ok=True)
    return d


# ─────────────────────────────────────────────────────────────────────────────
# STEP 1 — Upload custom PNG template
# POST /clubs/{club_id}/events/{event_id}/guest/template
# ─────────────────────────────────────────────────────────────────────────────

@router.post("/template")
async def upload_guest_template(
    club_id: PydanticObjectId,
    event_id: PydanticObjectId,
    file: UploadFile = File(...),
    _user: User = Depends(require_event_access),
):
    """Upload a PNG file as the guest's custom certificate background.

    The file is stored in storage/guest_templates/ and the path is saved
    on the Event document as ``guest_template_path``.
    """
    event = await Event.get(event_id)
    if not event or event.club_id != club_id:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Event not found")

    if not file.content_type or not file.content_type.startswith("image/"):
        raise HTTPException(
            status.HTTP_422_UNPROCESSABLE_ENTITY,
            "Only image files (PNG/JPG) are accepted as templates",
        )

    content = await file.read()
    if len(content) > 20 * 1024 * 1024:  # 20 MB hard cap
        raise HTTPException(
            status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
            "Template image must be smaller than 20 MB",
        )

    # Save with a unique name so concurrent uploads don't collide
    ext = Path(file.filename or "template.png").suffix or ".png"
    filename = f"guest_{club_id}_{event_id}_{uuid.uuid4().hex[:8]}{ext}"
    dest = _guest_templates_dir() / filename
    dest.write_bytes(content)

    await event.set({"guest_template_path": str(dest)})

    preview_url = f"/storage/guest_templates/{filename}"
    return {
        "message": "Template uploaded",
        "guest_template_path": str(dest),
        "preview_url": preview_url,
    }


# ─────────────────────────────────────────────────────────────────────────────
# STEP 2a — Upload XLSX, parse and return headers
# POST /clubs/{club_id}/events/{event_id}/guest/excel
# ─────────────────────────────────────────────────────────────────────────────

@router.post("/excel")
async def upload_guest_excel(
    club_id: PydanticObjectId,
    event_id: PydanticObjectId,
    file: UploadFile = File(...),
    _user: User = Depends(require_event_access),
):
    """Parse an uploaded XLSX file and return its column headers + row count.

    The parsed rows are stored temporarily on the Event document so that
    Step 4 (generation) can reference them without re-uploading.
    """
    event = await Event.get(event_id)
    if not event or event.club_id != club_id:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Event not found")

    content = await file.read()
    if len(content) > 20 * 1024 * 1024:
        raise HTTPException(
            status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
            "Excel file must be smaller than 20 MB",
        )

    try:
        from openpyxl import load_workbook
        wb = load_workbook(filename=io.BytesIO(content), read_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        headers_row = next(rows_iter, None)
        if not headers_row:
            raise HTTPException(status.HTTP_400_BAD_REQUEST, "Excel file is empty")

        headers = [str(h).strip() for h in headers_row if h is not None]
        if not headers:
            raise HTTPException(
                status.HTTP_400_BAD_REQUEST, "No column headers found in the first row"
            )

        # Parse all rows into list-of-dicts for later use
        parsed_rows: List[Dict] = []
        for row in rows_iter:
            if all(cell is None for cell in row):
                continue
            record: Dict = {}
            for col_idx, cell_value in enumerate(row):
                if col_idx < len(headers) and headers[col_idx]:
                    record[headers[col_idx]] = (
                        str(cell_value).strip() if cell_value is not None else ""
                    )
            parsed_rows.append(record)

        wb.close()
    except HTTPException:
        raise
    except Exception as exc:
        logger.error("Excel parse error: %s", exc)
        raise HTTPException(status.HTTP_400_BAD_REQUEST, f"Failed to parse Excel: {exc}")

    # Persist parsed data on the event (will be used in Step 4)
    await event.set({
        "guest_excel_data": parsed_rows,
        # Reset downstream state since a new file was uploaded
        "guest_selected_columns": None,
        "guest_email_column": None,
        "guest_generated_certs": None,
        "guest_emails_sent": False,
    })

    return {
        "headers": headers,
        "row_count": len(parsed_rows),
        "message": f"Parsed {len(parsed_rows)} data rows",
    }


# ─────────────────────────────────────────────────────────────────────────────
# STEP 2b — Save column selection + email column
# POST /clubs/{club_id}/events/{event_id}/guest/config
# ─────────────────────────────────────────────────────────────────────────────

class GuestConfigRequest(BaseModel):
    selected_columns: List[str]
    email_column: str


@router.post("/config")
async def save_guest_config(
    club_id: PydanticObjectId,
    event_id: PydanticObjectId,
    body: GuestConfigRequest,
    _user: User = Depends(require_event_access),
):
    """Persist which columns should be printed and which is the email column."""
    event = await Event.get(event_id)
    if not event or event.club_id != club_id:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Event not found")

    if not body.selected_columns:
        raise HTTPException(
            status.HTTP_422_UNPROCESSABLE_ENTITY,
            "At least one column must be selected",
        )
    if body.email_column not in body.selected_columns and body.email_column:
        # email_column doesn't have to be in selected_columns (it may be metadata-only)
        pass

    await event.set({
        "guest_selected_columns": body.selected_columns,
        "guest_email_column": body.email_column,
        # Reset certs so old stale certs are not mixed with new config
        "guest_generated_certs": None,
        "guest_emails_sent": False,
    })

    return {
        "message": "Column configuration saved",
        "selected_columns": body.selected_columns,
        "email_column": body.email_column,
    }


# ─────────────────────────────────────────────────────────────────────────────
# STEP 4 — Generate certificates (Pillow batch)
# POST /clubs/{club_id}/events/{event_id}/guest/generate
# ─────────────────────────────────────────────────────────────────────────────

@router.post("/generate")
async def generate_guest_certificates(
    club_id: PydanticObjectId,
    event_id: PydanticObjectId,
    _user: User = Depends(require_event_access),
):
    """Generate one PNG certificate per Excel row using the Pillow pipeline.

    Requires:
    - guest_template_path: set in Step 1
    - guest_excel_data:    set in Step 2a
    - guest_selected_columns + guest_email_column: set in Step 2b
    - FieldPosition confirmed:  set in Step 3 via the shared /field-positions endpoint

    Returns list of generated certificate paths and a summary.
    """
    event = await Event.get(event_id)
    if not event or event.club_id != club_id:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Event not found")

    # Validate prerequisites
    if not event.guest_template_path or not Path(event.guest_template_path).exists():
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            "No valid certificate template found. Please complete Step 1 first.",
        )
    if not event.guest_excel_data:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            "No Excel data found. Please complete Step 2 first.",
        )
    if not event.guest_selected_columns or not event.guest_email_column:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            "Column configuration not saved. Please complete Step 2 first.",
        )

    # Fetch field positions (Step 3) — we use cert_type="guest" for guest events
    fp = await FieldPosition.find_one(
        FieldPosition.event_id == event_id,
        FieldPosition.cert_type == "guest",
    )
    if not fp:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            "Field positions not configured. Please complete Step 3 first.",
        )

    template_path = Path(event.guest_template_path)
    output_dir = _guest_certs_dir(str(club_id), str(event_id))
    rows = event.guest_excel_data

    generated: List[str] = []
    errors: List[str] = []

    # Run CPU-bound Pillow work in a thread pool
    for idx, row in enumerate(rows):
        safe_name = f"cert_{idx+1:04d}_{uuid.uuid4().hex[:6]}.png"
        out_path = output_dir / safe_name
        try:
            await asyncio.to_thread(
                _render_guest_certificate,
                template_path,
                row,
                fp.column_positions,
                str(out_path),
            )
            generated.append(str(out_path))
        except Exception as exc:
            logger.error("Error generating cert for row %d: %s", idx + 1, exc)
            errors.append(f"Row {idx + 1}: {exc}")

    await event.set({
        "guest_generated_certs": generated,
        "guest_emails_sent": False,
    })

    return {
        "generated": len(generated),
        "errors": errors,
        "message": f"Generated {len(generated)} certificate(s).",
    }


def _render_guest_certificate(
    template_path: Path,
    fields: dict,
    column_positions: dict,
    output_path: str,
) -> None:
    """Synchronous Pillow rendering — called inside asyncio.to_thread."""
    from PIL import Image, ImageDraw, ImageFont

    _FONTS_DIR = Path(__file__).parent.parent / "static" / "fonts"
    _DEFAULT_FONT = _FONTS_DIR / "PlayfairDisplay.ttf"
    _ALT_FONT = _FONTS_DIR / "PlayfairDisplay.ttf"
    DEFAULT_FONT_PERCENT =2.7

    def _load_font(size: int, is_main: bool = True):
        try:
            f = _DEFAULT_FONT if is_main else _ALT_FONT
            if f.exists():
                return ImageFont.truetype(str(f), size)
        except Exception:
            pass
        return ImageFont.load_default()

    img = Image.open(str(template_path)).convert("RGBA")
    draw = ImageDraw.Draw(img)
    img_w, img_h = img.size

    for col_header, pos in (column_positions or {}).items():
        value = str((fields or {}).get(col_header, ""))
        if not value:
            continue
        fsp = float(pos.get("font_size_percent") or DEFAULT_FONT_PERCENT)
        fsp = max(1.0, min(fsp, 8.0))
        font_size = max(10, int(img_w * fsp / 100))
        font = _load_font(font_size)
        x = (pos["x_percent"] / 100) * img_w
        y = (pos["y_percent"] / 100) * img_h
        draw.text((x, y), value, font=font, fill=(30, 30, 30, 255), anchor="mm")

    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    if img.mode == "RGBA":
        bg = Image.new("RGB", img.size, "white")
        bg.paste(img, mask=img.split()[3])
        final = bg
    else:
        final = img.convert("RGB")
    final.save(str(out), format="PNG", optimize=False)


# ─────────────────────────────────────────────────────────────────────────────
# STEP 5a — Download all certificates as ZIP
# GET /clubs/{club_id}/events/{event_id}/guest/zip
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/zip")
async def download_guest_zip(
    club_id: PydanticObjectId,
    event_id: PydanticObjectId,
    _user: User = Depends(require_event_access),
):
    """Return all generated guest certificates as a single zip file."""
    event = await Event.get(event_id)
    if not event or event.club_id != club_id:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Event not found")

    if not event.guest_generated_certs:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            "No certificates have been generated yet. Please complete Step 4 first.",
        )

    existing = [p for p in event.guest_generated_certs if Path(p).exists()]
    if not existing:
        raise HTTPException(
            status.HTTP_404_NOT_FOUND,
            "Certificate files not found on disk. Please regenerate.",
        )

    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for path_str in existing:
            p = Path(path_str)
            zf.write(p, arcname=p.name)
    zip_buf.seek(0)

    event_name_safe = (event.name or "event").replace(" ", "_")
    return StreamingResponse(
        zip_buf,
        media_type="application/zip",
        headers={
            "Content-Disposition": f'attachment; filename="{event_name_safe}_certificates.zip"'
        },
    )


# ─────────────────────────────────────────────────────────────────────────────
# STEP 5b — Send emails
# POST /clubs/{club_id}/events/{event_id}/guest/send-emails
# ─────────────────────────────────────────────────────────────────────────────

@router.post("/send-emails")
async def send_guest_emails(
    club_id: PydanticObjectId,
    event_id: PydanticObjectId,
    _user: User = Depends(require_event_access),
):
    """Send each generated certificate to the email address in the email column.

    Matches generated certificate files (by index) to Excel rows.
    """
    event = await Event.get(event_id)
    if not event or event.club_id != club_id:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Event not found")

    if not event.guest_generated_certs:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            "No certificates available. Please complete Step 4 first.",
        )
    if not event.guest_email_column:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            "Email column not configured. Please complete Step 2 first.",
        )

    rows = event.guest_excel_data or []
    certs = event.guest_generated_certs or []
    email_col = event.guest_email_column

    # Try to get the club name for email body
    try:
        from ..models.club import Club
        club = await Club.get(club_id)
        club_name = club.name if club else "Your Club"
    except Exception:
        club_name = "Your Club"

    sent = 0
    failed = 0
    errors: List[str] = []

    for idx, (row, cert_path) in enumerate(zip(rows, certs)):
        recipient_email = row.get(email_col, "").strip()
        if not recipient_email:
            errors.append(f"Row {idx + 1}: No email address found in column '{email_col}'")
            failed += 1
            continue

        if not Path(cert_path).exists():
            errors.append(f"Row {idx + 1}: Certificate file not found — {cert_path}")
            failed += 1
            continue

        # Determine recipient name from common name columns
        recipient_name = ""
        for name_key in ("Name", "name", "Full Name", "full_name", "Student Name"):
            if name_key in row and row[name_key]:
                recipient_name = row[name_key]
                break
        if not recipient_name:
            recipient_name = recipient_email.split("@")[0]

        cert_number = f"GUEST-{str(event_id)[-6:]}-{idx+1:04d}"
        try:
            success = await send_certificate_email(
                recipient_email=recipient_email,
                recipient_name=recipient_name,
                cert_number=cert_number,
                event_name=event.name,
                club_name=club_name,
                png_path=cert_path,
            )
            if success:
                sent += 1
            else:
                failed += 1
                errors.append(f"Row {idx + 1}: Email delivery failed (rate limit or SMTP error)")
        except Exception as exc:
            failed += 1
            errors.append(f"Row {idx + 1}: {exc}")

    if sent > 0:
        await event.set({"guest_emails_sent": True})

    return {
        "sent": sent,
        "failed": failed,
        "errors": errors,
        "message": f"Sent {sent} email(s). {failed} failed.",
    }


# ─────────────────────────────────────────────────────────────────────────────
# STATUS — Poll current guest event state
# GET /clubs/{club_id}/events/{event_id}/guest/status
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/status")
async def get_guest_status(
    club_id: PydanticObjectId,
    event_id: PydanticObjectId,
    _user: User = Depends(require_event_access),
):
    """Return the current state of the guest wizard for this event."""
    event = await Event.get(event_id)
    if not event or event.club_id != club_id:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Event not found")

    # Check if field positions have been configured
    fp = await FieldPosition.find_one(
        FieldPosition.event_id == event_id,
        FieldPosition.cert_type == "guest",
    )

    template_url = None
    if event.guest_template_path:
        tmpl = Path(event.guest_template_path)
        if tmpl.exists():
            template_url = f"/storage/guest_templates/{tmpl.name}"

    return {
        "step1_complete": bool(event.guest_template_path and Path(event.guest_template_path).exists()),
        "template_url": template_url,
        "step2_complete": bool(event.guest_excel_data and event.guest_selected_columns and event.guest_email_column),
        "excel_row_count": len(event.guest_excel_data) if event.guest_excel_data else 0,
        "selected_columns": event.guest_selected_columns or [],
        "email_column": event.guest_email_column,
        "all_excel_headers": (
            list(event.guest_excel_data[0].keys()) if event.guest_excel_data else []
        ),
        "step3_complete": bool(fp),
        "field_positions": {
            "column_positions": fp.column_positions if fp else {},
            "display_width": fp.display_width if fp else 580.0,
        } if fp else None,
        "step4_complete": bool(event.guest_generated_certs),
        "generated_count": len(event.guest_generated_certs) if event.guest_generated_certs else 0,
        "step5_emails_sent": event.guest_emails_sent,
    }
